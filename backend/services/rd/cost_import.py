"""
BOM 成本 Excel 解析服务。

Excel 结构约定：
  Sheet 0（产品预估单）：汇总页，提取订单号
  其余 Sheet（跳过 Sheet1）：每个 Sheet = 一个 SKU
    第 1 行：表头（按列名匹配，而非固定位置）
    中间行：BOM 明细
    末行：合计（主件品号列为纯数字行数字符串）

BOM 列名（默认别名，可通过 SiteConfig cost_col_aliases 配置）：
  主件品号   → finished_code    品号        → parent_code
  元件品号   → child_code       元件品名    → child_name
  元件规格   → child_spec       元件品名分类→ child_category
  组成用量   → quantity         单价        → unit_price
  金额/总价  → total_price      标准号      → child_std_code
  序号       → seq
"""
import json
import re
from openpyxl import load_workbook
from database.base import db
from database.models.rd.cost import (
    CostSnapshot, CostSnapshotSku, CostBomNode, CostBomLine,
)
from database.models.product.finished import ProductFinished


# ── 汇总 Sheet 名关键词（不导入为 SKU 的 Sheet）────────────
_SKIP_SHEET_KEYWORDS = ('预估单', '预测单', '汇总', 'sheet1')

# ── 合计行识别：主件品号列为纯数字字符串（行数）──────────────
_ROW_COUNT_RE = re.compile(r'^\d+$')

# ── 默认列名别名表（可通过 SiteConfig 覆盖）────────────────
DEFAULT_COL_ALIASES = {
    'finished_code': ['主件品号', '成品品号'],
    'finished_name': ['主件品名', '成品品名'],
    'finished_spec': ['主件规格', '成品规格'],
    'parent_code':   ['品号'],
    'parent_name':   ['品名'],
    'parent_spec':   ['规格'],
    'seq':           ['序号'],
    'quantity':      ['组成用量', '用量'],
    'child_code':    ['元件品号', '子件品号'],
    'child_std_code':['标准号'],
    'child_name':    ['元件品名', '子件品名'],
    'child_spec':    ['元件规格', '子件规格'],
    'child_category':['元件品名分类', '品名分类', '分类'],
    'unit_price':    ['单价'],
    'total_price':   ['金额', '总价'],
}

# ── 当列名未匹配时的默认位置（备用）──────────────────────────
_DEFAULT_POSITIONS = {
    'finished_code': 0, 'finished_name': 1, 'finished_spec': 2,
    'parent_code': 3, 'parent_name': 4, 'parent_spec': 5,
    'seq': 6, 'quantity': 7,
    'child_code': 8, 'child_std_code': 9,
    'child_name': 10, 'child_spec': 11,
    'child_category': 12,
    'unit_price': 14, 'total_price': 15,
}


def load_col_aliases() -> dict:
    """从 SiteConfig 读取自定义列名别名；不存在则返回默认值。"""
    try:
        from database.models.account import SiteConfig
        cfg = SiteConfig.query.get('cost_col_aliases')
        if cfg and cfg.value:
            return json.loads(cfg.value)
    except Exception:
        pass
    return DEFAULT_COL_ALIASES


def _detect_col_map(header_row: list, aliases: dict = None) -> dict:
    """
    按列名（精确匹配）找到各字段在表头行中的列索引。
    未匹配到的字段使用 _DEFAULT_POSITIONS 中的备用位置。
    """
    if aliases is None:
        aliases = DEFAULT_COL_ALIASES

    col_map = dict(_DEFAULT_POSITIONS)                    # 先填入备用位置
    hdrs = [_str(c).strip() for c in header_row]         # 去两端空格，保留大小写

    for field, names in aliases.items():
        if field not in _DEFAULT_POSITIONS:
            continue
        for i, h in enumerate(hdrs):
            if h in names:                                # 精确匹配
                col_map[field] = i
                break

    return col_map


def _strip_version(code: str) -> str:
    """去掉品号版本后缀（-A01、-B02 等），保留基础码。"""
    if not code:
        return code
    return re.sub(r'-[A-Z]\d+$', '', code.strip())


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    f = _to_float(val)
    if f is None:
        return None
    return int(round(f))


def _str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _is_summary_sheet(name: str) -> bool:
    lower = name.lower()
    return any(kw in lower for kw in _SKIP_SHEET_KEYWORDS)


def _suggest_date(order_no: str | None) -> str | None:
    """从订单号中提取 YYYYMMDD 日期（如 2M2-SC20240620-050 → 2024-06-20）。"""
    if not order_no:
        return None
    m = re.search(r'(\d{4})(\d{2})(\d{2})', order_no)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y}-{mo:02d}-{d:02d}'
    return None


def _parse_summary_sheet(ws):
    """从汇总 Sheet 提取 order_no（订单号）和 SKU 列表的数量信息。"""
    order_no = None
    sku_qtys = {}  # finished_code -> qty_in_order

    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            s = _str(cell)
            if order_no is None:
                for kw in ('单号', '编号'):
                    if kw in s:
                        for sep in ('：', ':'):
                            if sep in s:
                                parts = s.split(sep, 1)
                                v = parts[1].strip() if len(parts) > 1 else ''
                                if v:
                                    order_no = v
                                    break
                        if not order_no:
                            for j in range(i + 1, len(row)):
                                v = _str(row[j])
                                if v:
                                    order_no = v
                                    break
                        if order_no:
                            break

        # 成品品号行：格式 品名|品号|单位|数量|备注|成本
        cells = [_str(c) for c in row]
        if len(cells) >= 4 and re.match(r'^1[2-9]\d{2}', cells[1]):
            code = cells[1]
            qty = _to_int(cells[3]) if len(cells) > 3 else None
            sku_qtys[code] = qty

    return order_no, sku_qtys


def _is_total_row(row_vals: list, col: dict) -> bool:
    """判断是否为合计行（末行）。"""
    # 合计行特征：物理第一列（通常为主件品号列）存放行数统计（纯整数）
    # 同时也检查 detected 的 finished_code 列，兼容列顺序不同的 Excel
    check_idxs = {0, col.get('finished_code', 0)}
    for idx in check_idxs:
        v = _str(row_vals[idx]) if len(row_vals) > idx else ''
        if _ROW_COUNT_RE.match(v):
            return True
    return False


def _parse_sku_sheet(ws, aliases: dict = None):
    """
    解析一个 SKU Sheet，返回：
      finished_code, finished_name, finished_spec,
      total_cost,
      lines: list of dict,
      col_map: dict（调试用）
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    # 从第 1 行（表头）按列名探测列位置
    col = _detect_col_map(rows[0], aliases)

    # 跳过第 1 行（表头）
    data_rows = rows[1:]

    finished_code = None
    finished_name = None
    finished_spec = None
    total_cost = None
    lines = []

    for row in data_rows:
        vals = list(row)
        # 补齐列数（防止短行索引越界）
        needed = max(col.values()) + 2 if col else 18
        while len(vals) < needed:
            vals.append(None)

        # 合计行
        if _is_total_row(vals, col):
            tc = _to_float(vals[col['total_price']]) if len(vals) > col['total_price'] else None
            if tc is None:
                tc = _to_float(vals[col['unit_price']]) if len(vals) > col['unit_price'] else None
            if tc is not None:
                total_cost = tc
            continue

        # 主件品号（第一次出现时记录成品信息）
        fc = _str(vals[col['finished_code']])
        if not fc:
            continue

        if finished_code is None:
            finished_code = fc
            finished_name = _str(vals[col['finished_name']])
            finished_spec = _str(vals[col['finished_spec']])

        parent_code = _str(vals[col['parent_code']])
        child_code  = _str(vals[col['child_code']])

        if not child_code:
            continue

        lines.append({
            'parent_code':    parent_code,
            'parent_name':    _str(vals[col['parent_name']]),
            'parent_spec':    _str(vals[col['parent_spec']]),
            'seq':            _to_int(vals[col['seq']]),
            'quantity':       _to_float(vals[col['quantity']]),
            'child_code':     child_code,
            'child_std_code': _str(vals[col['child_std_code']]),
            'child_name':     _str(vals[col['child_name']]),
            'child_spec':     _str(vals[col['child_spec']]),
            'child_category': _str(vals[col['child_category']]),
            'unit_price':     _to_float(vals[col['unit_price']]),
            'total_price':    _to_float(vals[col['total_price']]),
        })

    return {
        'finished_code': finished_code,
        'finished_name': finished_name,
        'finished_spec': finished_spec,
        'total_cost':    total_cost,
        'lines':         lines,
    }


def _upsert_node(code_with_ver: str, name: str, spec: str, category: str,
                 std_code: str, node_cache: dict) -> CostBomNode:
    """
    按去版本后缀的 code 查找或新建 CostBomNode。
    node_cache: {code -> CostBomNode} 本次导入会话内的缓存，避免重复查库。
    """
    base_code = _strip_version(code_with_ver)
    if not base_code:
        return None

    if base_code in node_cache:
        return node_cache[base_code]

    node = CostBomNode.query.filter_by(code=base_code).first()
    if node is None:
        node = CostBomNode(
            code=base_code,
            code_with_version=code_with_ver or None,
            name=name or None,
            spec=spec or None,
            category=category or None,
            std_code=std_code or None,
            node_type='material',
        )
        db.session.add(node)
    else:
        # 补充缺失字段（不覆盖已有值）
        if not node.code_with_version and code_with_ver:
            node.code_with_version = code_with_ver
        if not node.name and name:
            node.name = name
        if not node.spec and spec:
            node.spec = spec
        if not node.category and category:
            node.category = category
        if not node.std_code and std_code:
            node.std_code = std_code

    node_cache[base_code] = node
    return node


def preview_excel(file_path: str) -> dict:
    """
    仅解析 Excel，不写库，返回预览信息：
    { order_no, sku_count, skus: [{finished_code, finished_name, finished_spec, total_cost, lines}], warnings }
    """
    aliases = load_col_aliases()
    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    order_no = None
    sku_qtys = {}
    for name in sheet_names:
        if _is_summary_sheet(name):
            order_no, sku_qtys = _parse_summary_sheet(wb[name])
            break

    skus = []
    warnings = []
    for name in sheet_names:
        if _is_summary_sheet(name):
            continue
        ws = wb[name]
        parsed = _parse_sku_sheet(ws, aliases)
        if not parsed or not parsed['finished_code']:
            warnings.append(f'Sheet "{name}" 无法解析成品品号，已跳过')
            continue
        fc = parsed['finished_code']
        skus.append({
            'finished_code': fc,
            'finished_name': parsed['finished_name'] or '',
            'finished_spec': parsed['finished_spec'] or '',
            'total_cost':    parsed['total_cost'],
            'line_count':    len(parsed['lines']),
            'qty_in_order':  sku_qtys.get(fc),
            'lines':         parsed['lines'],
        })

    return {
        'order_no':       order_no or '',
        'suggested_date': _suggest_date(order_no),
        'sku_count':      len(skus),
        'skus':           skus,
        'warnings':       warnings,
    }


def import_excel(file_path: str, snapshot_date, notes: str, created_by: str) -> dict:
    """
    主入口：解析 Excel 并写入数据库。
    返回 { snapshot_id, sku_count, line_count, warnings }。
    """
    aliases = load_col_aliases()
    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    # 1. 读汇总 Sheet
    order_no = None
    sku_qtys = {}
    for name in sheet_names:
        if _is_summary_sheet(name):
            order_no, sku_qtys = _parse_summary_sheet(wb[name])
            break

    # 2. 创建快照
    snapshot = CostSnapshot(
        order_no=order_no or None,
        snapshot_date=snapshot_date or None,
        notes=notes or None,
        created_by=created_by or None,
    )
    db.session.add(snapshot)
    db.session.flush()

    # 3. 解析 SKU Sheets
    node_cache = {}
    sku_count = 0
    line_count = 0
    warnings = []
    all_parent_codes = set()
    sku_data_list = []

    for name in sheet_names:
        if _is_summary_sheet(name):
            continue
        ws = wb[name]
        parsed = _parse_sku_sheet(ws, aliases)
        if not parsed or not parsed['finished_code']:
            warnings.append(f'Sheet "{name}" 无法解析成品品号，已跳过')
            continue
        sku_data_list.append((name, parsed))
        for line in parsed['lines']:
            if line['parent_code']:
                all_parent_codes.add(_strip_version(line['parent_code']))

    for sheet_name, parsed in sku_data_list:
        fc = parsed['finished_code']

        pf = ProductFinished.query.filter_by(code=fc).first()
        pf_id = pf.id if pf else None
        qty = sku_qtys.get(fc)

        db_sku = CostSnapshotSku(
            snapshot_id=snapshot.id,
            finished_code=fc,
            finished_name=parsed['finished_name'] or None,
            finished_spec=parsed['finished_spec'] or None,
            product_finished_id=pf_id,
            qty_in_order=qty,
            total_cost=parsed['total_cost'],
        )
        db.session.add(db_sku)
        db.session.flush()

        # 确保成品本身有节点记录
        finished_base = _strip_version(fc)
        if finished_base not in node_cache:
            fn = CostBomNode.query.filter_by(code=finished_base).first()
            if fn is None:
                fn = CostBomNode(
                    code=finished_base,
                    code_with_version=fc,
                    name=parsed['finished_name'] or None,
                    spec=parsed['finished_spec'] or None,
                    node_type='finished',
                )
                db.session.add(fn)
            elif fn.node_type != 'finished':
                fn.node_type = 'finished'
            node_cache[finished_base] = fn
            db.session.flush()

        # 写 BOM 明细行
        for line in parsed['lines']:
            parent_base = _strip_version(line['parent_code'])
            child_base  = _strip_version(line['child_code'])

            parent_node = node_cache.get(parent_base)
            if parent_node is None:
                parent_node = _upsert_node(
                    line['parent_code'], line['parent_name'], line['parent_spec'],
                    '', '', node_cache,
                )
                db.session.flush()

            child_node = _upsert_node(
                line['child_code'], line['child_name'], line['child_spec'],
                line['child_category'], line['child_std_code'], node_cache,
            )
            if child_node is None:
                warnings.append(f'行 {line} 子件品号为空，已跳过')
                continue
            db.session.flush()

            db_line = CostBomLine(
                sku_id=db_sku.id,
                parent_node_id=parent_node.id,
                child_node_id=child_node.id,
                seq=line['seq'],
                quantity=line['quantity'],
                unit_price=line['unit_price'],
                total_price=line['total_price'],
            )
            db.session.add(db_line)
            line_count += 1

        sku_count += 1

    # 4. 后处理：将在其他行中作为父件出现的子件标记为 semi
    db.session.flush()
    for base_code, node in node_cache.items():
        if node.node_type == 'material' and base_code in all_parent_codes:
            node.node_type = 'semi'

    # 写入物料价格历史（bom_import）
    _write_material_prices(sku_data_list, node_cache, snapshot, snapshot_date, created_by,
                           from_parsed=True)

    db.session.commit()

    return {
        'snapshot_id': snapshot.id,
        'sku_count':   sku_count,
        'line_count':  line_count,
        'warnings':    warnings,
    }


def _write_material_prices(skus_data, node_cache, snapshot, snapshot_date, created_by,
                           from_parsed=False):
    """
    将 BOM 明细中的物料单价写入 cost_material_price（source=bom_import）。
    去重：同一次导入同一物料只写一条；相同 (node_id, price_date, unit_price) 已存在则跳过。
    from_parsed=True 时 skus_data 为 [(sheet_name, parsed_dict), ...]，否则为 [sku_data_dict, ...]
    """
    from database.models.rd.cost import CostMaterialPrice

    written_node_ids = set()  # 本次已处理的 node_id，每物料取首次价格

    # 查询当日已有的 bom_import 记录，用于跨次导入去重
    existing = set()
    if snapshot_date:
        rows = (
            CostMaterialPrice.query
            .filter_by(source='bom_import', price_date=snapshot_date)
            .with_entities(CostMaterialPrice.node_id, CostMaterialPrice.unit_price)
            .all()
        )
        existing = {(r.node_id, float(r.unit_price)) for r in rows}

    def iter_lines():
        if from_parsed:
            for _name, parsed in skus_data:
                yield from parsed.get('lines', [])
        else:
            for sku_data in skus_data:
                yield from sku_data.get('lines', [])

    for line in iter_lines():
        child_code = line.get('child_code', '')
        if not child_code:
            continue
        node = node_cache.get(_strip_version(child_code))
        if node is None:
            continue
        # 记录原材料，或外购半成品（整体采购，有直接单价）
        if node.node_type == 'material':
            pass
        elif node.node_type == 'semi' and node.is_purchased_semi:
            pass
        else:
            continue

        price = _to_float(line.get('unit_price'))
        if not price:
            continue

        if node.id in written_node_ids:
            continue
        written_node_ids.add(node.id)

        if (node.id, price) in existing:
            continue

        db.session.add(CostMaterialPrice(
            node_id=node.id,
            unit_price=price,
            price_date=snapshot_date,
            supplier_name=None,
            source='bom_import',
            snapshot_id=snapshot.id,
            created_by=created_by,
        ))


def import_from_data(data: dict, snapshot_date, notes: str, created_by: str) -> dict:
    """
    从预览数据（可能已由用户编辑）写入数据库，不重新解析 Excel。
    data 格式与 preview_excel() 返回值相同。
    """
    order_no  = data.get('order_no') or None
    skus_data = data.get('skus', [])
    # 外购半成品基础码集合（下级 BOM 行跳过不录入）
    purchased_base_codes = {_strip_version(c) for c in data.get('purchased_semi_codes', [])}

    snapshot = CostSnapshot(
        order_no=order_no,
        snapshot_date=snapshot_date or None,
        notes=notes or None,
        created_by=created_by or None,
    )
    db.session.add(snapshot)
    db.session.flush()

    node_cache = {}
    sku_count  = 0
    line_count = 0
    warnings   = []
    all_parent_codes = set()

    for sku_data in skus_data:
        for line in sku_data.get('lines', []):
            pc = line.get('parent_code', '')
            if pc and _strip_version(pc) not in purchased_base_codes:
                all_parent_codes.add(_strip_version(pc))

    for sku_data in skus_data:
        fc = sku_data.get('finished_code', '')
        if not fc:
            continue

        pf    = ProductFinished.query.filter_by(code=fc).first()
        pf_id = pf.id if pf else None

        db_sku = CostSnapshotSku(
            snapshot_id=snapshot.id,
            finished_code=fc,
            finished_name=sku_data.get('finished_name') or None,
            finished_spec=sku_data.get('finished_spec') or None,
            product_finished_id=pf_id,
            qty_in_order=_to_int(sku_data.get('qty_in_order')),
            total_cost=_to_float(sku_data.get('total_cost')),
        )
        db.session.add(db_sku)
        db.session.flush()

        finished_base = _strip_version(fc)
        if finished_base not in node_cache:
            fn = CostBomNode.query.filter_by(code=finished_base).first()
            if fn is None:
                fn = CostBomNode(
                    code=finished_base,
                    code_with_version=fc,
                    name=sku_data.get('finished_name') or None,
                    spec=sku_data.get('finished_spec') or None,
                    node_type='finished',
                )
                db.session.add(fn)
            elif fn.node_type != 'finished':
                fn.node_type = 'finished'
            node_cache[finished_base] = fn
            db.session.flush()

        for line in sku_data.get('lines', []):
            parent_base = _strip_version(line.get('parent_code', ''))

            # 外购半成品的下级不录入 BOM 行
            if parent_base in purchased_base_codes:
                continue

            parent_node = node_cache.get(parent_base)
            if parent_node is None:
                parent_node = _upsert_node(
                    line.get('parent_code', ''), line.get('parent_name', ''),
                    line.get('parent_spec', ''), '', '', node_cache,
                )
                db.session.flush()

            child_node = _upsert_node(
                line.get('child_code', ''), line.get('child_name', ''),
                line.get('child_spec', ''), line.get('child_category', ''),
                line.get('child_std_code', ''), node_cache,
            )
            if child_node is None:
                warnings.append('子件品号为空，已跳过')
                continue
            db.session.flush()

            db_line = CostBomLine(
                sku_id=db_sku.id,
                parent_node_id=parent_node.id,
                child_node_id=child_node.id,
                seq=_to_int(line.get('seq')),
                quantity=_to_float(line.get('quantity')),
                unit_price=_to_float(line.get('unit_price')),
                total_price=_to_float(line.get('total_price')),
            )
            db.session.add(db_line)
            line_count += 1

        sku_count += 1

    db.session.flush()
    purchased_base_codes = {_strip_version(c) for c in data.get('purchased_semi_codes', [])}
    for base_code, node in node_cache.items():
        if node.node_type == 'material' and base_code in all_parent_codes:
            node.node_type = 'semi'
    for base_code, node in node_cache.items():
        if node.node_type == 'semi' and base_code in purchased_base_codes:
            node.is_purchased_semi = True

    # 写入物料价格历史（bom_import）
    _write_material_prices(skus_data, node_cache, snapshot, snapshot_date, created_by)

    db.session.commit()

    return {
        'snapshot_id': snapshot.id,
        'sku_count':   sku_count,
        'line_count':  line_count,
        'warnings':    warnings,
    }
