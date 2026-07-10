import csv
import io
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import List, Dict
from database.repository.shipping import shipping_repository
from result import Result

from utils import now_cst

# 直辖市省名集合：这些省的 city 字段统一填写为省名本身
_MUNICIPALITY_PROVINCES = {'北京市', '天津市', '上海市', '重庆市'}

def _normalize_city(province: str, city) -> str:
    """直辖市 city 统一归一到省名（如北京市），避免出现区级名或 null"""
    if province in _MUNICIPALITY_PROVINCES:
        return province
    return city

# ── 必要列名（按列名匹配，与列顺序无关）─────────────
_REQUIRED_COL_NAMES = {
    '电商主订单号', '单据日期', '渠道名称', '渠道商', '渠道商名称',
    '最近操作人', '项次', '商品型号', '商品名称', '数量', '省份',
}

# 销退清单必要列名（与发货清单列名不同）
_REQUIRED_RETURN_COL_NAMES = {'平台订单', '交易日期', '品号', '数量', '仓库名称'}


def _build_col_map(header_row, required: set = None) -> Dict[str, int]:
    """
    扫描表头行，返回 {列名: 列索引} 的映射（去除首尾空格后匹配）。
    若必要列有缺失则抛出 ValueError，列出所有缺失列名。
    required 不传时使用发货清单的必要列集合。
    """
    if required is None:
        required = _REQUIRED_COL_NAMES
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            col_map[str(cell).strip()] = idx

    missing = required - col_map.keys()
    if missing:
        raise ValueError(
            '文件缺少以下必要列，请确认是否上传了正确的文件：\n'
            + '、'.join(sorted(missing))
        )
    return col_map


def _str(v) -> str:
    """安全转字符串，去除首尾空格，None 返回空字符串"""
    return str(v).strip() if v is not None else ''


def _parse_shipped_date(v):
    """解析单据日期：字符串 '2024/01/11' 或 datetime 对象"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_quantity(v):
    """解析数量，返回 Decimal 或 None"""
    if v is None:
        return None
    try:
        return Decimal(str(v).strip())
    except InvalidOperation:
        return None


def _get(row: tuple, idx: int):
    """安全取列值，超出范围返回 None"""
    return row[idx] if idx < len(row) else None


def _extract_row(row, col_map: Dict[str, int]) -> Dict:
    """按列名映射从行数据中提取所有字段"""
    def gc(name):  # get column value by name
        return _get(row, col_map[name]) if name in col_map else None

    return {
        'ecommerce_order_no': _str(gc('电商主订单号'))  or None,
        'line_no':            _str(gc('项次'))          or None,
        'shipped_date':       _parse_shipped_date(gc('单据日期')),
        'channel_name':       _str(gc('渠道名称'))      or None,
        'channel_code':       _str(gc('渠道商'))        or None,
        'channel_org_name':   _str(gc('渠道商名称'))    or None,
        'operator':           _str(gc('最近操作人'))    or None,
        'product_code':       _str(gc('商品型号'))      or None,
        'product_name':       _str(gc('商品名称'))      or None,
        'spec':               _str(gc('规格'))          or None,
        'quantity':           _parse_quantity(gc('数量')),
        'country':            _str(gc('国家'))          or None,
        'province':           _str(gc('省份'))          or None,
        'city':               _normalize_city(_str(gc('省份')) or '', _str(gc('市区')) or None),
        'district':           _str(gc('县区'))          or None,
        'street':             _str(gc('街道'))          or None,
        'address':            _str(gc('详细地址'))      or None,
        'buyer_remark':       _str(gc('买家留言'))      or None,
        'seller_remark':      _str(gc('商家备注'))      or None,
    }


def _parse_xlsx_rows(file_bytes: bytes) -> List[Dict]:
    """解析 xlsx：按列名匹配提取字段，与列顺序无关"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    col_map = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            col_map = _build_col_map(row)  # 校验必要列是否存在，返回列名→索引映射
            continue
        rows.append(_extract_row(row, col_map))
    wb.close()
    return rows


def _parse_csv_rows(file_bytes: bytes) -> List[Dict]:
    """解析 csv：按列名匹配提取字段，与列顺序无关"""
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows = []
    col_map = None
    for i, row in enumerate(reader):
        if i == 0:
            col_map = _build_col_map(row)  # 校验必要列是否存在，返回列名→索引映射
            continue
        rows.append(_extract_row(row, col_map))
    return rows


# ── 销退清单专用解析（列名：平台订单/交易日期/品号/数量）────────────

def _extract_return_row(row, col_map: Dict[str, int]) -> Dict:
    """按销退清单列名提取字段，仅提取 return_record 所需列"""
    def gc(name):
        return _get(row, col_map[name]) if name in col_map else None

    return {
        'ecommerce_order_no': _str(gc('平台订单'))  or None,
        'shipped_date':       _parse_shipped_date(gc('交易日期')),
        'product_code':       _str(gc('品号'))       or None,
        'quantity':           _parse_quantity(gc('数量')),
        'warehouse_name':     _str(gc('仓库名称'))   or None,
    }


def _parse_xlsx_return_rows(file_bytes: bytes) -> List[Dict]:
    """解析销退 xlsx：按销退列名匹配提取字段"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    col_map = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            col_map = _build_col_map(row, required=_REQUIRED_RETURN_COL_NAMES)
            continue
        rows.append(_extract_return_row(row, col_map))
    wb.close()
    return rows


def _parse_csv_return_rows(file_bytes: bytes) -> List[Dict]:
    """解析销退 csv：按销退列名匹配提取字段"""
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows = []
    col_map = None
    for i, row in enumerate(reader):
        if i == 0:
            col_map = _build_col_map(row, required=_REQUIRED_RETURN_COL_NAMES)
            continue
        rows.append(_extract_return_row(row, col_map))
    return rows


# ── 财务端专用解析（列名：平台订单/订单单号/交易日期/部门名称/品号/数量/省/市/区）────
_REQUIRED_FINANCE_COL_NAMES = {'交易日期', '部门名称', '品号', '数量', '省'}
_FINANCE_AFTERSALE_DEPT = '售后组'


def _extract_finance_order_no(row, col_map: Dict[str, int]) -> str:
    """
    补全平台订单号：
    1. 「平台订单」不为空 → 直接使用
    2. 「平台订单」为空但「订单单号」不为空 → 取「订单单号」中第一个「-」后的部分
    3. 两者都为空 → 返回空字符串（调用方跳过该行）
    """
    platform = _str(_get(row, col_map['平台订单']) if '平台订单' in col_map else None)
    if platform:
        return platform
    order_no = _str(_get(row, col_map['订单单号']) if '订单单号' in col_map else None)
    if order_no and '-' in order_no:
        return order_no.split('-', 1)[1]
    return ''


def _extract_finance_row(row, col_map: Dict[str, int]) -> Dict:
    """按财务清单列名提取字段"""
    def gc(name):
        return _get(row, col_map[name]) if name in col_map else None

    province = _str(gc('省')) or None
    return {
        'ecommerce_order_no': _extract_finance_order_no(row, col_map) or None,
        'channel_name':       _str(gc('部门名称'))  or None,
        'shipped_date':       _parse_shipped_date(gc('交易日期')),
        'product_code':       _str(gc('品号'))       or None,
        'product_name':       _str(gc('品名'))       or None,
        'spec':               _str(gc('规格'))       or None,
        'quantity':           _parse_quantity(gc('数量')),
        'province':           province,
        'city':               _normalize_city(province or '', _str(gc('市')) or None),
        'district':           _str(gc('区'))         or None,
    }


def _parse_xlsx_finance_rows(file_bytes: bytes):
    """解析财务 xlsx，返回 (shipping_rows, return_rows, aftersale_count)"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    shipping_rows, return_rows = [], []
    aftersale_count = 0
    col_map = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            col_map = _build_col_map(row, required=_REQUIRED_FINANCE_COL_NAMES)
            continue
        dept = _str(_get(row, col_map.get('部门名称', -1)))
        if dept == _FINANCE_AFTERSALE_DEPT:
            aftersale_count += 1
            continue
        r = _extract_finance_row(row, col_map)
        if not r.get('ecommerce_order_no'):
            continue
        qty = r.get('quantity')
        if qty is None:
            continue
        if qty > 0:
            shipping_rows.append(r)
        elif qty < 0:
            r['quantity'] = abs(qty)
            return_rows.append(r)
    wb.close()
    return shipping_rows, return_rows, aftersale_count


def _parse_csv_finance_rows(file_bytes: bytes):
    """解析财务 csv，返回 (shipping_rows, return_rows, aftersale_count)"""
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    shipping_rows, return_rows = [], []
    aftersale_count = 0
    col_map = None
    for i, row in enumerate(reader):
        if i == 0:
            col_map = _build_col_map(row, required=_REQUIRED_FINANCE_COL_NAMES)
            continue
        dept = _str(_get(row, col_map.get('部门名称', -1)))
        if dept == _FINANCE_AFTERSALE_DEPT:
            aftersale_count += 1
            continue
        r = _extract_finance_row(row, col_map)
        if not r.get('ecommerce_order_no'):
            continue
        qty = r.get('quantity')
        if qty is None:
            continue
        if qty > 0:
            shipping_rows.append(r)
        elif qty < 0:
            r['quantity'] = abs(qty)
            return_rows.append(r)
    return shipping_rows, return_rows, aftersale_count


def _merge_finance_shipping_rows(rows: List[Dict]):
    """财务发货行文件内合并：按 (order_no, product_code) 合并，无 line_no"""
    from collections import OrderedDict
    merged: OrderedDict = OrderedDict()
    merged_away: List[Dict] = []
    for row in rows:
        key = (row.get('ecommerce_order_no'), row.get('product_code'))
        if key in merged:
            existing_qty = merged[key].get('quantity') or Decimal('0')
            new_qty      = row.get('quantity') or Decimal('0')
            merged[key]['quantity'] = existing_qty + new_qty
            merged_away.append(row)
        else:
            merged[key] = dict(row)
    return list(merged.values()), merged_away


def _serialize_finance_skipped_row(r: Dict) -> Dict:
    """财务导入跳过行序列化"""
    shipped = r.get('shipped_date')
    qty = r.get('quantity')
    return {
        'ecommerce_order_no': r.get('ecommerce_order_no'),
        'product_code':       r.get('product_code'),
        'shipped_date':       shipped.strftime('%Y-%m-%d') if shipped else None,
        'quantity':           float(qty) if qty is not None else None,
    }


def _serialize_row(r: Dict) -> Dict:
    """将内存中的原始行转为可 JSON 序列化的 dict（全列）"""
    qty     = r.get('quantity')
    shipped = r.get('shipped_date')
    return {
        'ecommerce_order_no': r.get('ecommerce_order_no'),
        'line_no':            r.get('line_no'),
        'shipped_date':       shipped.strftime('%Y-%m-%d') if shipped else None,
        'channel_name':       r.get('channel_name'),
        'channel_code':       r.get('channel_code'),
        'channel_org_name':   r.get('channel_org_name'),
        'operator':           r.get('operator'),
        'product_code':       r.get('product_code'),
        'product_name':       r.get('product_name'),
        'spec':               r.get('spec'),
        'quantity':           float(qty) if qty is not None else None,
        'country':            r.get('country'),
        'province':           r.get('province'),
        'city':               r.get('city'),
        'district':           r.get('district'),
        'street':             r.get('street'),
        'address':            r.get('address'),
        'buyer_remark':       r.get('buyer_remark'),
        'seller_remark':      r.get('seller_remark'),
    }


def _serialize_return_row(r: Dict) -> Dict:
    """将销退行转为可 JSON 序列化的 dict"""
    qty     = r.get('quantity')
    shipped = r.get('shipped_date')
    return {
        'ecommerce_order_no': r.get('ecommerce_order_no'),
        'shipped_date':       shipped.strftime('%Y-%m-%d') if shipped else None,
        'product_code':       r.get('product_code'),
        'quantity':           float(qty) if qty is not None else None,
        'warehouse_name':     r.get('warehouse_name'),
    }


def _merge_rows(rows: List[Dict]):
    """
    将文件内相同 (ecommerce_order_no, line_no, product_code) 的行合并：
    - 数量（quantity）累加，其余字段取首行的值
    - 返回 (merged_rows, merged_away_rows)
    """
    from collections import OrderedDict
    merged:      OrderedDict = OrderedDict()
    merged_away: List[Dict]  = []
    for row in rows:
        key = (row.get('ecommerce_order_no'), row.get('line_no'), row.get('product_code'))
        if key in merged:
            existing_qty = merged[key].get('quantity') or Decimal('0')
            new_qty      = row.get('quantity') or Decimal('0')
            merged[key]['quantity'] = existing_qty + new_qty
            merged_away.append(row)   # 记录被合并掉的原始行
        else:
            merged[key] = dict(row)
    return list(merged.values()), merged_away


def _merge_return_rows(rows: List[Dict]):
    """
    将文件内相同 (ecommerce_order_no, product_code, shipped_date) 的销退行合并：
    - 数量（quantity）累加，warehouse_name 取首行的值
    - 返回 (merged_rows, merged_away_rows)
    """
    from collections import OrderedDict
    merged:      OrderedDict = OrderedDict()
    merged_away: List[Dict]  = []
    for row in rows:
        key = (row.get('ecommerce_order_no'), row.get('product_code'), row.get('shipped_date'))
        if key in merged:
            existing_qty = merged[key].get('quantity') or Decimal('0')
            new_qty      = row.get('quantity') or Decimal('0')
            merged[key]['quantity'] = existing_qty + new_qty
            merged_away.append(row)
        else:
            merged[key] = dict(row)
    return list(merged.values()), merged_away


def _resolve_orders(order_nos: List[str], source: str = 'shipping', progress_cb=None):
    """
    对给定订单号列表，执行成品组合匹配，写入 shipping_order_finished。
    source: 'shipping' 或 'finance'，决定从哪个来源的 shipping_record 读取产成品数据。
    匹配逻辑：
      1. 取订单内所有产成品编码及数量
      2. 遍历产品库中所有成品，找出成品的产成品集合是订单产成品集合的子集
      3. 贪心匹配：优先匹配产成品数最多的成品，按最小分量计算成品数量
    """
    if not order_nos:
        return

    from database.models.product.finished import ProductFinished, ProductPackaged
    from database.base import db

    total_orders = len(order_nos)

    if progress_cb:
        progress_cb('preparing', message='正在加载产品库…', total=total_orders)

    # 加载所有成品及其产成品关联（selectinload 避免 N+1）
    from sqlalchemy.orm import selectinload
    from database.models.product.finished import PackagedEquivalent
    from collections import defaultdict
    finished_list = ProductFinished.query.options(
        selectinload(ProductFinished.packaged_list)
    ).all()
    # finished_map: finished_code → (finished_name, frozenset of packaged codes)
    finished_map = {}
    for f in finished_list:
        packaged_codes = frozenset(p.code for p in f.packaged_list)
        if packaged_codes:
            finished_map[f.code] = (f.code, _get_finished_name(f), packaged_codes)

    # 按产成品数量从多到少排序（贪心：优先匹配更复杂的成品）
    sorted_finished = sorted(
        finished_map.values(),
        key=lambda x: len(x[2]),
        reverse=True,
    )

    # 加载通用件等效映射：{code: set(含自身及所有等效码)}
    equiv_map: Dict = defaultdict(set)
    for ep in PackagedEquivalent.query.all():
        equiv_map[ep.code_a].update([ep.code_a, ep.code_b])
        equiv_map[ep.code_b].update([ep.code_a, ep.code_b])

    def _avail(code, remaining):
        """该槽位可用数量：有精确码则只算精确码，缺货才回退到等效码之和"""
        exact = remaining.get(code, 0)
        if exact > 0:
            return exact
        return sum(remaining.get(c, 0) for c in equiv_map.get(code, {code}))

    def _consume(code, qty, remaining):
        """消耗产成品：优先消耗 code 自身，再消耗等效码（字典序）"""
        left = qty
        for c in sorted(equiv_map.get(code, {code}), key=lambda x: (x != code, x)):
            if left <= 0:
                break
            avail = remaining.get(c, 0)
            if avail <= 0:
                continue
            take = min(avail, left)
            remaining[c] -= take
            left -= take
            if remaining[c] <= 0:
                del remaining[c]

    # 分批加载发货数据，每批推送一次进度
    CHUNK = 2000
    order_data: Dict = {}
    for i in range(0, total_orders, CHUNK):
        chunk = order_nos[i:i + CHUNK]
        order_data.update(shipping_repository.get_order_products(chunk, source=source))
        if progress_cb:
            progress_cb('preparing', message='正在加载发货数据…',
                        current=min(i + CHUNK, total_orders), total=total_orders)

    # 分批加载销退数据
    return_data: Dict = {}
    for i in range(0, total_orders, CHUNK):
        chunk = order_nos[i:i + CHUNK]
        return_data.update(shipping_repository.get_order_return_products(chunk))
        if progress_cb:
            progress_cb('preparing', message='正在加载销退数据…',
                        current=min(i + CHUNK, total_orders), total=total_orders)

    # 对每个订单预先跑一次贪心匹配，得到 {order_no: {finished_code: return_qty}}
    return_resolved: Dict[str, Dict] = {}
    for order_no, return_products in return_data.items():
        remaining_ret = dict(return_products)  # {product_code: abs_qty}
        order_ret = {}
        for f_code, f_name, required_codes in sorted_finished:
            if not all(_avail(code, remaining_ret) > 0 for code in required_codes):
                continue
            min_qty = min(_avail(code, remaining_ret) for code in required_codes)
            if min_qty <= 0:
                continue
            for code in required_codes:
                _consume(code, min_qty, remaining_ret)
            order_ret[f_code] = order_ret.get(f_code, 0) + min_qty
        return_resolved[order_no] = order_ret

    # 清除旧结果（只清除同一 source 的记录）
    shipping_repository.delete_order_finished(order_nos, source=source)

    resolved_at = now_cst()
    to_insert = []
    total_orders = len(order_data)

    for idx, (order_no, data) in enumerate(order_data.items()):
        # 每处理 100 个订单推送一次进度
        if progress_cb and idx % 100 == 0:
            progress_cb('resolving', current=idx, total=total_orders)
        remaining = dict(data['product_codes'])  # {product_code: qty}
        meta = data['meta']
        order_ret = return_resolved.get(order_no, {})
        matched_any = False

        for f_code, f_name, required_codes in sorted_finished:
            # 检查订单中是否有该成品所需的全部产成品（含等效码）
            if not all(_avail(code, remaining) > 0 for code in required_codes):
                continue
            # 可组合的数量 = 各槽位可用数量（含等效码）中最小的那个
            min_qty = min(_avail(code, remaining) for code in required_codes)
            if min_qty <= 0:
                continue
            # 扣减已使用的产成品数量（优先消耗原码，再消耗等效码）
            for code in required_codes:
                _consume(code, min_qty, remaining)
            rq = Decimal(str(order_ret.get(f_code, 0)))
            sq = Decimal(str(min_qty))
            to_insert.append({
                'ecommerce_order_no': order_no,
                'finished_code':      f_code,
                'finished_name':      f_name,
                'quantity':           sq,
                'return_quantity':    rq,
                'actual_quantity':    sq - rq,
                'shipped_date':       meta['shipped_date'],
                'operator':           meta['operator'],
                'channel_name':       meta['channel_name'],
                'channel_code':       meta.get('channel_code'),
                'channel_org_name':   meta.get('channel_org_name'),
                'province':           meta['province'],
                'city':               _normalize_city(meta['province'], meta.get('city')),
                'district':           meta.get('district'),
                'source':             source,
                'resolved_at':        resolved_at,
            })
            matched_any = True

        # 剩余未匹配的产成品：写一行 finished_code=None 的记录，方便追踪
        for code, qty in remaining.items():
            if qty > 0:
                sq = Decimal(str(qty))
                to_insert.append({
                    'ecommerce_order_no': order_no,
                    'finished_code':      None,
                    'finished_name':      None,
                    'quantity':           sq,
                    'return_quantity':    Decimal('0'),
                    'actual_quantity':    sq,
                    'shipped_date':       meta['shipped_date'],
                    'operator':           meta['operator'],
                    'channel_name':       meta['channel_name'],
                    'channel_code':       meta.get('channel_code'),
                    'channel_org_name':   meta.get('channel_org_name'),
                    'province':           meta['province'],
                    'city':               _normalize_city(meta['province'], meta.get('city')),
                    'district':           meta.get('district'),
                    'source':             source,
                    'resolved_at':        resolved_at,
                })

    if progress_cb:
        progress_cb('resolving', current=total_orders, total=total_orders)

    shipping_repository.bulk_insert_order_finished(to_insert, progress_cb=progress_cb)


def _get_finished_name(finished) -> str:
    """取成品显示名称：优先用中文名，无则用编码"""
    if finished.model and finished.model.name:
        return finished.model.name
    return finished.code


class ShippingService:

    def import_shipping(self, filename: str, file_bytes: bytes,
                        progress_cb=None, cancel_check=None) -> Dict:
        """导入发货清单：文件内合并 → 与库去重 → 插入新行 → 成品组合"""
        def notify(step, **kwargs):
            if progress_cb:
                progress_cb(step, **kwargs)

        batch = None
        try:
            notify('parsing')
            name_lower = filename.lower()
            rows = _parse_csv_rows(file_bytes) if name_lower.endswith('.csv') else _parse_xlsx_rows(file_bytes)

            total = len(rows)
            rows, merged_away_rows = _merge_rows(rows)   # 文件内相同 key 合并

            # 与数据库比对，分出新行和跳过行
            keys          = [(r.get('ecommerce_order_no'), r.get('line_no'), r.get('product_code')) for r in rows]
            existing_keys = shipping_repository.get_existing_keys(keys, record_type='shipping')
            new_rows      = [r for r in rows if (r.get('ecommerce_order_no'), r.get('line_no'), r.get('product_code')) not in existing_keys]
            skipped_rows  = [r for r in rows if (r.get('ecommerce_order_no'), r.get('line_no'), r.get('product_code')) in existing_keys]
            notify('parsed', total=total)

            imported_at = now_cst()
            notify('inserting', current=0, total=len(new_rows))
            batch = shipping_repository.create_batch('shipping', filename, total, imported_at)

            def on_insert_progress(current, total_rows):
                if cancel_check and cancel_check():
                    raise InterruptedError('用户已中止导入')
                notify('inserting', current=current, total=total_rows)

            inserted = shipping_repository.bulk_insert_shipping(batch.id, new_rows, progress_cb=on_insert_progress,
                                                                    record_type='shipping')
            notify('inserted', inserted=inserted, skipped=len(skipped_rows))

            # 对本批次新增的订单触发成品组合
            new_order_nos = shipping_repository.get_new_order_nos_by_source(batch.id, source='shipping')
            if new_order_nos:
                notify('resolving', current=0, total=len(new_order_nos))
                _resolve_orders(new_order_nos, source='shipping', progress_cb=progress_cb)

            return {
                'total':             total,
                'merged_away':       len(merged_away_rows),
                'inserted':          inserted,
                'skipped':           len(skipped_rows),
                'skipped_rows':      [_serialize_row(r) for r in skipped_rows],
                'merged_away_rows':  [_serialize_row(r) for r in merged_away_rows],
            }
        except Exception:
            # 先清理可能存在的脏事务，再删除批次数据
            if batch is not None:
                try:
                    from database.base import db
                    db.session.rollback()
                    shipping_repository.delete_batch(batch.id)
                except Exception:
                    pass
            raise

    def import_return(self, filename: str, file_bytes: bytes,
                      progress_cb=None, cancel_check=None) -> Dict:
        """导入销退清单：过滤排除仓库 → 提取负数量行 → 匹配订单 → 文件内合并 → 与库去重 → 插入 return_record → 重算成品组合"""
        def notify(step, **kwargs):
            if progress_cb:
                progress_cb(step, **kwargs)

        batch = None
        try:
            notify('parsing')
            name_lower = filename.lower()
            all_rows = _parse_csv_return_rows(file_bytes) if name_lower.endswith('.csv') else _parse_xlsx_return_rows(file_bytes)

            total = len(all_rows)

            # 仅处理数量为负数的行（不在导入时过滤仓库，保存全量数据）
            negative_rows = [r for r in all_rows if r.get('quantity') is not None and r.get('quantity') < 0]

            # 按订单号匹配：仅保留发货库中存在对应订单的行
            candidate_order_nos = list({r.get('ecommerce_order_no') for r in negative_rows
                                        if r.get('ecommerce_order_no')})
            existing_order_set = shipping_repository.get_existing_order_nos(candidate_order_nos)

            matched_rows   = [r for r in negative_rows if r.get('ecommerce_order_no') in existing_order_set]
            unmatched_rows = [r for r in negative_rows if r.get('ecommerce_order_no') not in existing_order_set]

            notify('parsed', total=total)

            # 文件内相同 key (order_no, product_code, shipped_date) 合并
            matched_rows, merged_away_rows = _merge_return_rows(matched_rows)

            # 与数据库比对去重（检查 return_record 已有记录）
            keys          = [(r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) for r in matched_rows]
            existing_keys = shipping_repository.get_existing_return_keys(keys)
            new_rows      = [r for r in matched_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) not in existing_keys]
            skipped_rows  = [r for r in matched_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) in existing_keys]

            imported_at = now_cst()
            notify('inserting', current=0, total=len(new_rows))
            batch = shipping_repository.create_batch('return', filename, total, imported_at)

            def on_insert_progress(current, total_rows):
                if cancel_check and cancel_check():
                    raise InterruptedError('用户已中止导入')
                notify('inserting', current=current, total=total_rows)

            inserted = shipping_repository.bulk_insert_return(batch.id, new_rows,
                                                              progress_cb=on_insert_progress)
            notify('inserted', inserted=inserted, skipped=len(skipped_rows))

            # 对受影响的订单重新计算成品组合（含销退数量）
            affected_order_nos = shipping_repository.get_return_affected_order_nos(batch.id)
            if affected_order_nos:
                notify('resolving', current=0, total=len(affected_order_nos))
                _resolve_orders(affected_order_nos, progress_cb=progress_cb)

            return {
                'total':             total,
                'negative_count':    len(negative_rows),
                'unmatched':         len(unmatched_rows),
                'inserted':          inserted,
                'skipped':           len(skipped_rows),
                'merged_away':       len(merged_away_rows),
                'skipped_rows':      [_serialize_return_row(r) for r in skipped_rows],
                'merged_away_rows':  [_serialize_return_row(r) for r in merged_away_rows],
                'unmatched_rows':    [_serialize_return_row(r) for r in unmatched_rows],
            }
        except Exception:
            if batch is not None:
                try:
                    from database.base import db
                    db.session.rollback()
                    shipping_repository.delete_batch(batch.id)
                except Exception:
                    pass
            raise

    def import_finance(self, filename: str, file_bytes: bytes,
                       progress_cb=None, cancel_check=None) -> Dict:
        """导入财务清单：过滤售后组 → 拆分发货/销退 → 文件内合并 → 与库去重 → 插入 → 成品组合"""
        def notify(step, **kwargs):
            if progress_cb:
                progress_cb(step, **kwargs)

        batch = None
        try:
            notify('parsing')
            name_lower = filename.lower()
            if name_lower.endswith('.csv'):
                shipping_rows, return_rows, aftersale_count = _parse_csv_finance_rows(file_bytes)
            else:
                shipping_rows, return_rows, aftersale_count = _parse_xlsx_finance_rows(file_bytes)

            total = len(shipping_rows) + len(return_rows) + aftersale_count

            # 文件内合并
            shipping_rows, shipping_merged_away = _merge_finance_shipping_rows(shipping_rows)
            return_rows,   return_merged_away   = _merge_return_rows(return_rows)

            # DB 去重：发货行按 (order_no, product_code, date, source='finance')
            s_keys          = [(r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) for r in shipping_rows]
            existing_s_keys = shipping_repository.get_existing_finance_keys(s_keys)
            new_shipping    = [r for r in shipping_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) not in existing_s_keys]
            skipped_shipping= [r for r in shipping_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) in existing_s_keys]

            # DB 去重：销退行（复用 return_record 去重）
            r_keys          = [(r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) for r in return_rows]
            existing_r_keys = shipping_repository.get_existing_return_keys(r_keys)
            new_returns     = [r for r in return_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) not in existing_r_keys]
            skipped_returns = [r for r in return_rows if (r.get('ecommerce_order_no'), r.get('product_code'), r.get('shipped_date')) in existing_r_keys]

            notify('parsed', total=total)

            imported_at = now_cst()
            notify('inserting', current=0, total=len(new_shipping) + len(new_returns))
            batch = shipping_repository.create_batch('finance', filename, total, imported_at)

            def on_insert_progress(current, total_rows):
                if cancel_check and cancel_check():
                    raise InterruptedError('用户已中止导入')
                notify('inserting', current=current, total=total_rows)

            inserted_shipping = shipping_repository.bulk_insert_shipping(
                batch.id, new_shipping,
                progress_cb=on_insert_progress,
                record_type='shipping', source='finance',
            )
            inserted_returns = shipping_repository.bulk_insert_return(
                batch.id, new_returns,
                progress_cb=on_insert_progress,
            )
            notify('inserted',
                   inserted=inserted_shipping,
                   inserted_returns=inserted_returns,
                   skipped=len(skipped_shipping),
                   aftersale_filtered=aftersale_count)

            # 对本批次财务来源新增订单触发成品组合
            new_order_nos = shipping_repository.get_new_order_nos_by_source(batch.id, source='finance')
            if new_order_nos:
                notify('resolving', current=0, total=len(new_order_nos))
                _resolve_orders(new_order_nos, source='finance', progress_cb=progress_cb)

            return {
                'total':              total,
                'aftersale_filtered': aftersale_count,
                'inserted':           inserted_shipping,
                'inserted_returns':   inserted_returns,
                'skipped':            len(skipped_shipping),
                'skipped_returns':    len(skipped_returns),
                'skipped_rows':       [_serialize_finance_skipped_row(r) for r in skipped_shipping],
            }
        except Exception:
            if batch is not None:
                try:
                    from database.base import db
                    db.session.rollback()
                    shipping_repository.delete_batch(batch.id)
                except Exception:
                    pass
            raise

    def get_operators(self) -> List[Dict]:
        return shipping_repository.get_all_operators()

    def classify_operators(self, items: List[Dict]) -> Dict:
        count = shipping_repository.classify_operators(items)
        return {'updated': count}

    def resolve_stale(self) -> Dict:
        """刷新所有 is_stale 的订单组合（按 source 分别处理）"""
        stale_pairs = shipping_repository.get_stale_order_nos()  # [(order_no, source), ...]
        # 按 source 分组
        from collections import defaultdict
        by_source = defaultdict(list)
        for order_no, src in stale_pairs:
            by_source[src].append(order_no)
        for src, order_nos in by_source.items():
            _resolve_orders(order_nos, source=src)
        return {'resolved': len(stale_pairs)}

    def get_stats(self) -> Dict:
        return shipping_repository.get_stats()

    def resolve_all(self, progress_cb=None) -> Dict:
        """全量重新计算所有订单的成品组合（发货端 + 财务端分别 resolve）"""
        shipping_nos = shipping_repository.get_all_order_nos_by_source('shipping')
        finance_nos  = shipping_repository.get_all_order_nos_by_source('finance')
        total = len(shipping_nos) + len(finance_nos)
        if shipping_nos:
            _resolve_orders(shipping_nos, source='shipping', progress_cb=progress_cb)
        if finance_nos:
            _resolve_orders(finance_nos, source='finance', progress_cb=progress_cb)
        return {'resolved': total}

    def get_shipped_dates(self) -> List[str]:
        return shipping_repository.get_distinct_shipped_dates()

    def get_warehouses(self) -> List[Dict]:
        return shipping_repository.get_all_warehouses()

    def save_warehouse_filters(self, items: List[Dict]) -> Dict:
        count = shipping_repository.save_warehouse_filters(items)
        return {'updated': count}

    def get_chart_options(self, date_start=None, date_end=None, source: str = 'shipping') -> Dict:
        return shipping_repository.get_chart_options(date_start=date_start, date_end=date_end, source=source)

    def get_chart_data(self, params: Dict) -> Dict:
        return shipping_repository.get_chart_data(params)

    def get_product_monthly(self, code: str) -> Result:
        data = shipping_repository.get_product_monthly(code)
        return Result.ok(data)

    def get_orders(self, page: int, size: int, filters: Dict, sort_field: str, sort_order: str) -> Dict:
        return shipping_repository.get_orders(page, size, filters, sort_field, sort_order)


shipping_service = ShippingService()
