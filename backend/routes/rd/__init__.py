from flask import Blueprint, request, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io, urllib.parse, os, re

rd_bp = Blueprint('rd', __name__)

# ── 样式常量 ──────────────────────────────────────────

_THIN  = Side(style='thin')
_THICK = Side(style='medium')

def _border(left=True, right=True, top=True, bottom=True, thick=False):
    s = _THICK if thick else _THIN
    return Border(
        left=s   if left   else Side(),
        right=s  if right  else Side(),
        top=s    if top    else Side(),
        bottom=s if bottom else Side(),
    )

def _font(name='宋体', size=10, bold=False, color='000000'):
    return Font(name=name, size=size, bold=bold, color=color)

def _align(horiz='left', vert='center', wrap=False):
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

FILL_HEADER = _fill('F0EDE6')
FILL_WHITE  = _fill('FFFFFF')
FILL_CANCEL = _fill('FDECEA')   # 取消行：浅红
BORDER_ALL  = _border()


def _set(ws, row, col, value, font=None, align=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fill:   cell.fill      = fill
    return cell


def _merge(ws, r1, c1, r2, c2, value, font=None, align=None, fill=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    top_left = ws.cell(row=r1, column=c1, value=value)
    if font:  top_left.font      = font
    if align: top_left.alignment = align
    if fill:  top_left.fill      = fill
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL
    return top_left


# ── checkbox 文本 ──────────────────────────────────────

def _auto_col_widths(ws, min_w=5, max_w=50):
    """根据单元格内容自动调整列宽（CJK 字符计 2，ASCII 计 1）。
    合并单元格按列数均分宽度，多行内容取最长行。"""
    # 找出所有合并区域的顶左格 → 列 span；及非顶左格（跳过）
    merge_spans = {}
    merged_secondary = set()
    for mc in ws.merged_cells.ranges:
        merge_spans[(mc.min_row, mc.min_col)] = mc.max_col - mc.min_col + 1
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                if r != mc.min_row or c != mc.min_col:
                    merged_secondary.add((r, c))

    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in merged_secondary or cell.value is None:
                continue
            text = str(cell.value)
            # 取多行中最长一行的宽度
            line_w = max(
                sum(2 if ord(ch) > 0x2E7F else 1 for ch in line)
                for line in text.split('\n')
            ) + 2  # +2 留白
            # 合并单元格按列数均分
            span = merge_spans.get((cell.row, cell.column), 1)
            per_col = line_w / span if span > 1 else line_w
            col = cell.column
            if per_col > col_max.get(col, 0):
                col_max[col] = per_col

    for col in range(1, ws.max_column + 1):
        w = col_max.get(col, min_w)
        ws.column_dimensions[get_column_letter(col)].width = max(min(round(w), max_w), min_w)


def _mark(options, selected, other_text=''):
    """生成 ☑/☐ 选择文本；other_text 为"其他"被选中时的自定义说明（显示为 其他：XXX）"""
    parts = []
    for opt in options:
        checked = opt in (selected if isinstance(selected, list) else [selected])
        if checked and opt == '其他' and other_text:
            parts.append(f'☑ 其他：{other_text}')
        else:
            parts.append(f'{"☑" if checked else "☐"} {opt}')
    return '    '.join(parts)


# ── BOM 比对 ──────────────────────────────────────────

def _level_str(v):
    """将层次单元格值统一转为字符串"""
    if v is None:
        return ''
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _level_sort_key(level_str):
    """将 '1.1.2' 转为可排序元组 (1, 1, 2)"""
    try:
        return tuple(int(x) for x in level_str.split('.') if x)
    except ValueError:
        return (0,)


_BOM_REQUIRED_COLS = ['层次', '图号', '品名', '规格', '数量', '单位', '状态']

def _validate_bom(path, role='before'):
    """校验 BOM 文件合法性，返回错误消息字符串；无误返回 None。
    role='any'   ：只校验列名，不校验状态
    role='before'：变更前文件，要求状态列全部为「已发布」（已废弃，保留供兼容）
    role='after' ：变更审核中文件，要求状态列不能全是「已发布」
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        label = '变更前文件' if role == 'before' else '变更审核中文件'
        return f'{label}无法读取：{e}'
    ws = wb.active

    # 1. 列名核验
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None:
            col_map[str(h).strip()] = c
    missing = [col for col in _BOM_REQUIRED_COLS if col not in col_map]
    label = '变更前文件' if role == 'before' else '变更审核中文件'
    if missing:
        wb.close()
        return f'{label}缺少必要列：{", ".join(missing)}'

    # 2. 状态列校验（跳过空行）
    status_col = col_map['状态']
    statuses = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, status_col).value
        if v is not None and str(v).strip():
            statuses.append(str(v).strip())
    wb.close()

    if not statuses:
        return f'{label}中未找到任何有效数据行'

    if role == 'after':
        if all(s == '已发布' for s in statuses):
            return '变更审核中文件的状态列全部为「已发布」，该文件应包含处于审核中状态的物料'

    return None


def _parse_bom(path):
    """读取并清洗 BOM 文件，返回 {(层次, 编码): item} 字典"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 表头 → 列号映射
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None:
            col_map[str(h).strip()] = c

    def _get(row, name, default=''):
        c = col_map.get(name)
        return ws.cell(row, c).value if c else default

    items = {}
    for r in range(2, ws.max_row + 1):
        level   = _level_str(_get(r, '层次'))
        drawing = str(_get(r, '图号') or '').strip()

        # 清洗：跳过无层次或无图号的行
        if not level or not drawing:
            continue

        # 以最后一个"-"分割编码和版本
        idx = drawing.rfind('-')
        code    = drawing[:idx]  if idx > 0 else drawing
        version = drawing[idx+1:] if idx > 0 else ''

        # 清洗：跳过编码为空的行（图号格式异常）
        if not code:
            continue

        # 清洗：过滤编码含"."的项、以"14ST10"开头的项
        if '.' in code or code.startswith('14ST10'):
            continue

        qty_raw = _get(r, '数量', 1)
        try:
            qty = float(qty_raw)
        except (TypeError, ValueError):
            qty = 1.0

        raw_spec = str(_get(r, '规格') or '').strip()
        items[(level, code)] = {
            'level':   level,
            'code':    code,
            'version': version,
            'drawing': drawing,
            'name':    str(_get(r, '品名') or '').strip(),
            'spec':    _complete_spec_version(raw_spec, version),
            'qty':     qty,
            'unit':    str(_get(r, '单位') or 'PCS').strip(),
            'status':  str(_get(r, '状态') or '').strip(),
        }
    return items


def _qty_str(qty):
    return str(int(qty)) if qty == int(qty) else str(qty)


# ── 版本号推算 ──────────────────────────────────────────

_VERSION_RE = re.compile(r'^([A-Za-z]+)(\d+)$')

def _complete_spec_version(spec, version):
    """规格版本补全：若图号版本为 'D01'，而规格中含 '_D'（只有字母、缺数字），
    自动补全为 '_D01'。避免误替换后面已有数字的情况。"""
    if not spec or not version:
        return spec
    m = _VERSION_RE.match(version.strip())
    if not m:
        return spec
    letters = m.group(1)
    num     = m.group(2)
    # 匹配 _{letters}（不区分大小写），且后面不跟数字
    pattern = re.compile(r'(_' + re.escape(letters) + r')(?!\d)', re.IGNORECASE)
    return pattern.sub(lambda x: x.group(1) + num, spec)


def _next_version(version, is_non_common):
    """推算变更后版本号
    - 通用变更审核中（is_non_common=False）：数字+1，字母不变。A01→A02，B02→B03
    - 非通用变更审核中（is_non_common=True）：最后一个字母+1，数字重置01。A01→B01，B02→C01
    """
    m = _VERSION_RE.match(version.strip())
    if not m:
        return version  # 无法解析，保持原样
    letters   = m.group(1).upper()
    number    = int(m.group(2))
    num_width = len(m.group(2))  # 保持数字位宽（如 01 → 宽度 2）
    if is_non_common:
        new_letters = letters[:-1] + chr(ord(letters[-1]) + 1)
        new_number  = 1
    else:
        new_letters = letters
        new_number  = number + 1
    return f"{new_letters}{str(new_number).zfill(num_width)}"


def _derive_new_drawing(item):
    """根据 item 的状态推算新图号；若非审核中则返回原图号"""
    status = item.get('status', '')
    if status not in ('通用变更审核中', '非通用变更审核中'):
        return item['drawing']
    is_non_common = (status == '非通用变更审核中')
    new_ver = _next_version(item['version'], is_non_common)
    return f"{item['code']}-{new_ver}"


def _update_spec_version(spec, old_ver, new_ver):
    """将规格字符串中旧版本号替换为新版本号（处理 -版本 和 _版本 两种前缀格式）"""
    if not spec or not old_ver or not new_ver or old_ver == new_ver:
        return spec
    result = spec.replace(f'-{old_ver}', f'-{new_ver}')
    result = result.replace(f'_{old_ver}', f'_{new_ver}')
    return result


def _compare_bom(before_path, after_path):
    """比对两个 BOM，返回 {changes, stats}

    规则：
    - 状态为"审核中"的物料 → 生成取消行（旧版本）+ 新增行（推算后的新版本）
    - 纯新增物料 → 生成一行新增（若为审核中则图号推算为新版本）
    - 仅在 before 存在的物料 → 生成一行取消
    - 排序按层次深度优先（1 < 1.1 < 1.1.1 < 1.2 < 2 ...）
    """
    before = _parse_bom(before_path)
    after  = _parse_bom(after_path)

    # 层次 → item 映射（层次在同一份 BOM 中唯一），用于查找父级
    before_by_level = {item['level']: item for item in before.values()}
    after_by_level  = {item['level']: item for item in after.values()}

    def _parent_item(level, by_level):
        """返回父级 item 或 None"""
        idx = level.rfind('.')
        if idx < 0:
            return None
        return by_level.get(level[:idx])

    def _old_main_drawing(level):
        """取消行/deleted行的主件图号：after BOM 中父级的当前图号（旧版本）"""
        p = _parent_item(level, after_by_level)
        return p['drawing'] if p else ''

    def _new_main_drawing(level):
        """新增行/added行的主件图号：after BOM 中父级的推算新图号（若父级也在变更）"""
        p = _parent_item(level, after_by_level)
        return _derive_new_drawing(p) if p else ''

    def _before_main_drawing(level):
        """deleted 行主件图号：before BOM 中父级的图号"""
        p = _parent_item(level, before_by_level)
        return p['drawing'] if p else ''

    # 按层次排序（深度优先）
    all_keys = sorted(
        set(before) | set(after),
        key=lambda k: _level_sort_key(k[0])
    )

    changes   = []
    n_version = 0   # 版本变更物料对数
    n_added   = 0
    n_deleted = 0

    for key in all_keys:
        in_b, in_a = key in before, key in after

        if in_b and in_a:
            # 在两个 BOM 中都存在：检查是否为审核中（即将变更）
            a = after[key]
            b = before[key]
            is_ecr = a['status'] in ('通用变更审核中', '非通用变更审核中')
            qty_changed = abs(b['qty'] - a['qty']) > 1e-9

            if not (is_ecr or qty_changed):
                continue  # 无变化，跳过

            n_version += 1

            if is_ecr:
                # 版本变更（含通用/非通用）：生成取消+新增两行
                is_non_common = (a['status'] == '非通用变更审核中')
                kind = '非通用变更' if is_non_common else '通用变更'

                old_drawing = a['drawing']
                old_version = a['version']
                new_version = _next_version(old_version, is_non_common)
                new_drawing = f"{a['code']}-{new_version}"
                new_spec    = _update_spec_version(a['spec'], old_version, new_version)

                # 取消行：取消旧版本，主件图号为 after 父级当前图号
                changes.append({
                    'row_type':      'cancel',
                    'change_kind':   kind,
                    'level':         key[0],
                    'main_drawing':  _old_main_drawing(a['level']),
                    'drawing':       old_drawing,
                    'name':          a['name'],
                    'spec':          a['spec'],
                    'qty':           b['qty'],
                    'change_method': '取消',
                })
                # 新增行：新增新版本，主件图号为 after 父级推算后的新图号
                changes.append({
                    'row_type':      'add',
                    'change_kind':   kind,
                    'level':         key[0],
                    'main_drawing':  _new_main_drawing(a['level']),
                    'drawing':       new_drawing,
                    'name':          a['name'],
                    'spec':          new_spec,
                    'qty':           a['qty'],
                    'change_method': '新增',
                })
            else:
                # 纯数量变更：只生成单行，变更方式为"数量变更"，取替代关系显示数量变化
                qty_desc = f"{_qty_str(b['qty'])}→{_qty_str(a['qty'])} {a['unit']}"
                changes.append({
                    'row_type':      'added',
                    'change_kind':   '数量变更',
                    'qty_desc':      qty_desc,
                    'level':         key[0],
                    'main_drawing':  _new_main_drawing(a['level']),
                    'drawing':       a['drawing'],
                    'name':          a['name'],
                    'spec':          a['spec'],
                    'qty':           a['qty'],
                    'change_method': '数量变更',
                })

        elif in_b:
            # 仅在 before 中存在（已删除）
            b = before[key]
            n_deleted += 1
            changes.append({
                'row_type':      'deleted',
                'change_kind':   '删除',
                'level':         key[0],
                'main_drawing':  _before_main_drawing(b['level']),
                'drawing':       b['drawing'],
                'name':          b['name'],
                'spec':          b['spec'],
                'qty':           b['qty'],
                'change_method': '取消',
            })

        else:
            # 仅在 after 中存在（新增）
            a = after[key]
            n_added += 1
            is_ecr_new    = a['status'] in ('通用变更审核中', '非通用变更审核中')
            is_non_common = (a['status'] == '非通用变更审核中')
            old_version   = a['version']
            new_version   = _next_version(old_version, is_non_common) if is_ecr_new else old_version
            new_drawing   = f"{a['code']}-{new_version}" if is_ecr_new else a['drawing']
            new_spec      = _update_spec_version(a['spec'], old_version, new_version)
            changes.append({
                'row_type':      'added',
                'change_kind':   '新增',
                'level':         key[0],
                'main_drawing':  _new_main_drawing(a['level']),
                'drawing':       new_drawing,
                'name':          a['name'],
                'spec':          new_spec,
                'qty':           a['qty'],
                'change_method': '新增',
            })

    for i, ch in enumerate(changes):
        ch['seq'] = i + 1

    stats = {
        'version': n_version,
        'added':   n_added,
        'deleted': n_deleted,
        'total':   len(changes),
    }
    return {'changes': changes, 'stats': stats}


# ── ECN 选项常量 ──────────────────────────────────────

_IMPORT_OPTS         = ['立即导入', '清化库存', '随单导入']
_AFFECTED_FILE_OPTS  = ['图纸', '模具', 'QC检验图', '包装标准', '作业指导书', '材料比较单', 'BOM', '品检规范', '其他']


# ── 生成 ECR xlsx ──────────────────────────────────────

def _build_ecr_xlsx(d: dict, changes=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '变更申请单'

    # 列宽在内容写完后统一自动调整（见文件末尾 _auto_col_widths 调用）

    # 行高（固定行）
    for r, h in {1:14, 2:26, 3:20, 4:18, 5:18, 6:16, 7:36, 8:16, 9:56, 10:16, 11:16, 12:16}.items():
        ws.row_dimensions[r].height = h

    f_normal = _font()
    f_bold   = _font(bold=True)
    f_title  = _font(size=16, bold=True)
    f_small  = _font(size=9)
    f_th     = _font(size=9, bold=True)
    f_detail = _font(size=9)
    a_center = _align('center')
    a_left   = _align('left')
    a_left_w = _align('left', wrap=True)
    a_right  = _align('right')

    # 第 1 行：文件编号
    _merge(ws, 1, 1, 1, 18, '2M2-QM-25-01-A1', font=f_small, align=a_right)
    for c in range(1, 19):
        ws.cell(1, c).border = Border()

    # 第 2 行：标题
    _merge(ws, 2, 1, 2, 18, '变 更 申 请 单', font=f_title, align=a_center)
    for c in range(1, 19):
        ws.cell(2, c).border = Border()

    # 第 3 行：基本信息
    issuing_unit       = d.get('issuing_unit', '')
    change_type        = d.get('change_type', '')
    change_type_custom = d.get('change_type_custom', '')
    change_type_text   = _mark(['设计变更', '制程变更', '其他'], change_type, change_type_custom)

    # 发出单位值缩为 1 列，其余左移一格，变更类型值扩至 4 列（14-17）
    _merge(ws, 3, 1,  3, 1,  '发出单位', font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 2,  3, 2,  issuing_unit, font=f_normal, align=a_left)
    _merge(ws, 3, 3,  3, 3,  '日期',      font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 4,  3, 5,  d.get('date', ''), font=f_normal, align=a_center)
    _merge(ws, 3, 6,  3, 6,  '变更编码',  font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 7,  3, 8,  d.get('ecr_code', ''), font=f_normal, align=a_center)
    _merge(ws, 3, 9,  3, 9,  '变更项目',  font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 10, 3, 13, d.get('project', ''), font=f_normal, align=a_left)
    _merge(ws, 3, 14, 3, 14, '变更类型',  font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 15, 3, 18, change_type_text, font=f_normal, align=a_left_w)

    # 第 4 行：分发单位
    dist_opts = ['研发', '业务', '采购', '生产', '生管', '品牌', '服务', '品管']
    _merge(ws, 4, 1, 4, 2,  '分发单位', font=f_bold,  align=a_center, fill=FILL_HEADER)
    _merge(ws, 4, 3, 4, 18, _mark(dist_opts, d.get('distribution', [])), font=f_normal, align=a_left)

    # 第 5 行：变更原因
    reason_opts          = ['品质不良', '价格变动', '设计优化', '结构优化', '成本优化', '工艺优化', '其他']
    change_reason        = d.get('change_reason', '')
    change_reason_custom = d.get('change_reason_custom', '')
    reason_text          = _mark(reason_opts, change_reason, change_reason_custom)
    _merge(ws, 5, 1, 5, 2,  '变更原因', font=f_bold,  align=a_center, fill=FILL_HEADER)
    _merge(ws, 5, 3, 5, 18, reason_text, font=f_normal, align=a_left)

    # 第 6-9 行：变更主题 + 变更内容说明
    _merge(ws, 6, 1, 6, 18, '变更主题：', font=f_bold, align=a_left, fill=FILL_HEADER)
    _merge(ws, 7, 1, 7, 18, d.get('change_subject', ''), font=f_normal, align=a_left_w)
    _merge(ws, 8, 1, 8, 18, '变更内容说明（详细说明更改哪些内容）：', font=f_bold, align=a_left, fill=FILL_HEADER)
    _merge(ws, 9, 1, 9, 18, d.get('change_desc', ''), font=f_normal, align=a_left_w)

    # 第 10 行：变更明细标签
    _merge(ws, 10, 1, 10, 18, '变更明细：', font=f_bold, align=a_left, fill=FILL_HEADER)

    # 第 11-12 行：明细表头
    def th(r1, c1, r2, c2, val):
        _merge(ws, r1, c1, r2, c2, val, font=f_th, align=a_center, fill=FILL_HEADER)

    # 列布局（共18列）：序1 主件图号2 图号3 层次4 品名5-6 规格7-8 变更方式9 取替代10 各部门11-16 处置方式17 责任人18
    th(11, 1,  12, 1,  '序号')
    th(11, 2,  12, 2,  '主件图号')
    th(11, 3,  12, 3,  '图号')
    th(11, 4,  12, 4,  '层次')
    th(11, 5,  12, 6,  '品名')
    th(11, 7,  12, 8,  '规格')
    th(11, 9,  12, 9,  '变更方式')
    th(11, 10, 12, 10, '取替代关系\n（原材料、半成品）')
    _merge(ws, 11, 11, 11, 16, '各部门问题反馈', font=f_th, align=a_center, fill=FILL_HEADER)
    for label, col in [('研发',11),('采购',12),('品管',13),('生管',14),('生产',15),('服务',16)]:
        _merge(ws, 12, col, 12, col, label, font=f_th, align=a_center, fill=FILL_HEADER)
    th(11, 17, 12, 17, '处置方式')
    th(11, 18, 12, 18, '责任人')

    # 第 13+ 行：明细数据（有比对结果则填充，否则留 6 行空行）
    detail_data = changes or []
    n_rows      = max(len(detail_data), 6)

    # 预扫描：找出 cancel+add 对，记录 {cancel行索引: change_kind}
    rd_merge_map = {}   # cancel_i -> change_kind
    rd_skip_set  = set()  # add_i（取替代关系列已被合并，跳过）
    j = 0
    while j < len(detail_data):
        if (detail_data[j].get('row_type') == 'cancel'
                and j + 1 < len(detail_data)
                and detail_data[j + 1].get('row_type') == 'add'):
            rd_merge_map[j] = detail_data[j].get('change_kind', '')
            rd_skip_set.add(j + 1)
            j += 2
        else:
            j += 1

    for i in range(n_rows):
        r = 13 + i
        ws.row_dimensions[r].height = 18
        for c in range(1, 19):
            ws.cell(row=r, column=c).border = BORDER_ALL

        if i < len(detail_data):
            ch       = detail_data[i]
            row_type = ch.get('row_type', '')
            row_fill = FILL_CANCEL if row_type in ('cancel', 'deleted') else None

            def _dc(col, val, al=None, _r=r, _fill=row_fill):
                cell = ws.cell(row=_r, column=col, value=val)
                cell.font      = f_detail
                cell.alignment = al or a_center
                if _fill:
                    cell.fill = _fill
                return cell

            # 取消行：对所有单元格（含合并区域内的空格）统一着色
            if row_fill:
                for c in range(1, 19):
                    ws.cell(row=r, column=c).fill = row_fill

            _dc(1, ch.get('seq', i + 1))
            _dc(2, ch.get('main_drawing', ''))        # 主件图号
            _dc(3, ch.get('drawing',      ''))        # 图号
            _dc(4, ch.get('level',        ''))        # 层次
            _merge(ws, r, 5, r, 6, ch.get('name', ''), font=f_detail, align=a_left, fill=row_fill)   # 品名（5-6合并）
            _merge(ws, r, 7, r, 8, ch.get('spec', ''), font=f_detail, align=a_left, fill=row_fill)   # 规格（7-8合并）
            _dc(9, ch.get('change_method', ''), a_left)

            # 取替代关系列（col 10）：cancel+add 对合并两行并填变更类型；单行填 qty_desc 或 change_kind
            if i in rd_merge_map:
                kind_text = rd_merge_map[i]
                _merge(ws, r, 10, r + 1, 10, kind_text, font=f_detail, align=a_center)
            elif i not in rd_skip_set:
                _dc(10, ch.get('qty_desc') or ch.get('change_kind', ''))

    # 填写人员行（紧接明细末尾）
    submitter_row = 13 + n_rows
    ws.row_dimensions[submitter_row].height = 16
    submitter = d.get('submitter', issuing_unit)
    _merge(ws, submitter_row, 1, submitter_row, 18,
           f'填写人员：{submitter}', font=f_normal, align=a_right)

    # Logo（左上角，锚定 A1）
    _logo_path = os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '..', '..', '..', 'src', 'assets', 'logo-banner.png'
    ))
    if os.path.exists(_logo_path):
        try:
            from openpyxl.drawing.image import Image as XlImage
            img = XlImage(_logo_path)
            img.width  = 110
            img.height = 30
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            _EPX = 9525   # EMU per pixel (96 DPI)
            marker = AnchorMarker(col=0, colOff=5 * _EPX, row=0, rowOff=3 * _EPX)
            img.anchor = OneCellAnchor(_from=marker,
                                       ext=XDRPositiveSize2D(cx=img.width * _EPX, cy=img.height * _EPX))
            ws.add_image(img)
        except Exception:
            pass

    # 根据实际内容自动调整列宽
    _auto_col_widths(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 生成 ECN xlsx ──────────────────────────────────────

def _build_ecn_xlsx(d: dict, changes=None) -> bytes:
    """生成变更通知单 xlsx
    列布局（12列 A-L）：
    A(1) : 标签列（信息行）/ 序号（明细行）
    B-C(2-3) : 主件图号（合并）
    D-E(4-5) : 图号（合并）
    F-G(6-7) : 品名（合并）
    H-I(8-9) : 规格（合并）
    J(10): 处理意见（导入方式选项）
    K(11): 负责人
    L(12): 备注
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '变更通知单'
    NC = 15  # 共15列：序号|主件图号(2)|图号(2)|层次|品名(2)|规格(2)|变更方式|取替代关系|处理意见|负责人|备注

    f_normal = _font()
    f_bold   = _font(bold=True)
    f_title  = _font(size=16, bold=True)
    f_small  = _font(size=9)
    f_th     = _font(size=9, bold=True)
    f_detail = _font(size=9)
    a_center = _align('center')
    a_left   = _align('left')
    a_left_w = _align('left', wrap=True)
    a_right  = _align('right')

    # 行高
    for r, h in {1:14, 2:26, 3:18, 4:18, 5:18, 6:18, 7:48, 8:36, 9:24}.items():
        ws.row_dimensions[r].height = h

    # 第1行：文件编号
    _merge(ws, 1, 1, 1, NC, '2M2-QM-25-03-A2', font=f_small, align=a_right)
    for c in range(1, NC + 1):
        ws.cell(1, c).border = Border()

    # 第2行：标题
    _merge(ws, 2, 1, 2, NC, '变 更 通 知 单', font=f_title, align=a_center)
    for c in range(1, NC + 1):
        ws.cell(2, c).border = Border()

    # 第3行：基本信息
    issuing_unit = d.get('issuing_unit', '')
    product      = d.get('product', d.get('project', ''))
    date_str     = d.get('date', '')
    ecn_code     = d.get('ecn_code', '')
    responsible  = d.get('responsible', '')

    # 列分配（共15列）：发出单位(1)|值(2-3)|产品型号(4)|值(5-7)|负责人(8)|值(9-10)|日期(11)|值(12)|ECN编号(13-15)
    _merge(ws, 3, 1,  3, 1,  '发出单位', font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 2,  3, 3,  issuing_unit, font=f_normal, align=a_left)
    _merge(ws, 3, 4,  3, 4,  '产品型号', font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 5,  3, 7,  product,      font=f_normal, align=a_left)
    _merge(ws, 3, 8,  3, 8,  '负责人',   font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 9,  3, 10, responsible,  font=f_normal, align=a_left)
    _merge(ws, 3, 11, 3, 11, '日期',     font=f_bold,   align=a_center, fill=FILL_HEADER)
    _merge(ws, 3, 12, 3, 12, date_str,    font=f_normal, align=a_center)
    _merge(ws, 3, 13, 3, NC, ecn_code,    font=f_bold,   align=a_center, fill=FILL_HEADER)

    # 通用信息行：标签列(1) + 内容列(2-NC)
    def _info_row(r, label, value, wrap=False):
        _merge(ws, r, 1, r, 1, label, font=f_bold, align=a_center, fill=FILL_HEADER)
        _merge(ws, r, 2, r, NC, value,
               font=f_normal, align=(_align('left', wrap=True) if wrap else a_left))

    import_meth   = d.get('import_method', '')
    distribution  = d.get('distribution', [])
    change_reason = d.get('change_reason', '')
    cr_custom     = d.get('change_reason_custom', '')
    affected      = d.get('affected_files', [])

    # 第4行：导入方式
    _info_row(4, '导入方式', _mark(_IMPORT_OPTS, import_meth))
    # 第5行：分发单位
    dist_opts = ['研发', '业务', '采购', '生产', '生管', '品牌', '服务', '品管']
    _info_row(5, '分发单位', _mark(dist_opts, distribution))
    # 第6行：变更原因
    reason_opts = ['品质不良', '价格变动', '设计优化', '结构优化', '成本优化', '工艺优化', '其他']
    _info_row(6, '变更原因', _mark(reason_opts, change_reason, cr_custom))
    # 第7行：变更内容说明
    _info_row(7, '变更内容', d.get('change_desc', ''), wrap=True)
    # 第8行：影响文件
    _info_row(8, '影响文件', _mark(_AFFECTED_FILE_OPTS, affected), wrap=True)

    responsible = d.get('responsible', '')

    # 第9行：明细表头
    def _th(c1, c2, val):
        _merge(ws, 9, c1, 9, c2, val, font=f_th, align=a_center, fill=FILL_HEADER)

    _th(1,  1,  '序号')
    _th(2,  3,  '主件图号')
    _th(4,  5,  '图号')
    _th(6,  6,  '层次')
    _th(7,  8,  '品名')
    _th(9,  10, '规格')
    _th(11, 11, '变更方式')
    _th(12, 12, '取替代关系')
    _th(13, 13, '处理意见')
    _th(14, 14, '负责人')
    _th(15, 15, '备注')

    # 明细数据行
    detail_data = changes or []
    n_rows      = max(len(detail_data), 5)

    # 预扫描：找出 cancel+add 对，取替代关系列需跨两行合并
    sub_merge_map = {}   # cancel_i -> 取替代关系文本
    sub_skip_set  = set()  # add_i（已被合并，跳过写入）
    j = 0
    while j < len(detail_data):
        if (detail_data[j].get('row_type') == 'cancel'
                and j + 1 < len(detail_data)
                and detail_data[j + 1].get('row_type') == 'add'):
            ch_j = detail_data[j]
            sub_merge_map[j] = ch_j.get('substitution') or ch_j.get('qty_desc') or ch_j.get('change_kind', '')
            sub_skip_set.add(j + 1)
            j += 2
        else:
            j += 1

    for i in range(n_rows):
        r = 10 + i
        ws.row_dimensions[r].height = 18
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL

        if i >= len(detail_data):
            continue

        ch       = detail_data[i]
        row_type = ch.get('row_type', '')
        row_fill = FILL_CANCEL if row_type in ('cancel', 'deleted') else None

        if row_fill:
            for c in range(1, NC + 1):
                ws.cell(row=r, column=c).fill = row_fill

        def _dc(col, val, al=None, _r=r, _fill=row_fill):
            cell = ws.cell(row=_r, column=col, value=val)
            cell.font      = f_detail
            cell.alignment = al or a_center
            if _fill:
                cell.fill = _fill
            return cell

        _dc(1, ch.get('seq', i + 1))
        _merge(ws, r, 2,  r, 3,  ch.get('main_drawing', ''), font=f_detail, align=a_center, fill=row_fill)
        _merge(ws, r, 4,  r, 5,  ch.get('drawing', ''),      font=f_detail, align=a_center, fill=row_fill)
        _dc(6, ch.get('level', ''), a_center)                             # 层次
        _merge(ws, r, 7,  r, 8,  ch.get('name', ''),         font=f_detail, align=a_left,   fill=row_fill)
        _merge(ws, r, 9,  r, 10, ch.get('spec', ''),         font=f_detail, align=a_left,   fill=row_fill)
        _dc(11, ch.get('change_method', ''), a_center)                    # 变更方式
        # 取替代关系（col 12）：cancel+add 对合并两行；单行直接写
        if i in sub_merge_map:
            _merge(ws, r, 12, r + 1, 12, sub_merge_map[i], font=f_detail, align=a_center)
        elif i not in sub_skip_set:
            _dc(12, ch.get('substitution') or ch.get('qty_desc') or ch.get('change_kind', ''), a_center)
        _dc(13, ch.get('handling', ''), a_center)                         # 处理意见
        _dc(14, ch.get('responsible_person', ''), a_center)               # 负责人
        _dc(15, '', a_left)                                                # 备注（空白）

    # ── 页脚区（按图片布局）─────────────────────────────
    # 列分区：col1-2=标签 | col3-4=生产 | col5-6=服务 | col7-9=品管 | col10-11=生管 | col12-13=研发 | col14-15=空余
    footer = 10 + n_rows

    f_dept  = _font(size=9, bold=True)
    a_ctr_w = _align('center', wrap=True)

    # ── 行 1/2：品管追踪记录（标签跨2行）+ 部门名称行 ──
    qt = footer
    ws.row_dimensions[qt].height   = 20
    ws.row_dimensions[qt + 1].height = 20

    # 标签：跨2行、2列
    _merge(ws, qt, 1, qt + 1, 2, '品管\n追踪记录', font=f_dept, align=a_ctr_w, fill=FILL_HEADER)

    # 部门名称（第1行）
    _merge(ws, qt, 3,  qt, 4,  '生  产', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, qt, 5,  qt, 6,  '服  务', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, qt, 7,  qt, 9,  '品  管', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, qt, 10, qt, 11, '生  管', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, qt, 12, qt, 13, '研  发', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, qt, 14, qt, NC, '',        font=f_normal, align=a_center)

    # 签名空白（第2行，对应各部门）
    for c1, c2 in [(3, 4), (5, 6), (10, 11), (12, 13), (14, NC)]:
        _merge(ws, qt + 1, c1, qt + 1, c2, '', font=f_normal, align=a_center)
    _merge(ws, qt + 1, 7, qt + 1, 9, '', font=f_normal, align=a_center)

    # ── 行 3：确认日期 ──
    cd = qt + 2
    ws.row_dimensions[cd].height = 18
    _merge(ws, cd, 1,  cd, 2,  '确认日期', font=f_dept, align=a_center, fill=FILL_HEADER)
    for c1, c2 in [(3, 4), (5, 6), (10, 11), (12, 13), (14, NC)]:
        _merge(ws, cd, c1, cd, c2, '', font=f_normal, align=a_center)
    _merge(ws, cd, 7, cd, 9, '', font=f_normal, align=a_center)

    # ── 行 4：备注 ──
    nr = cd + 1
    ws.row_dimensions[nr].height = 28
    _merge(ws, nr, 1,  nr, 2,  '备  注', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, nr, 3,  nr, NC, '',        font=f_normal, align=a_left)

    # ── 行 5：结案 ──
    ar = nr + 1
    ws.row_dimensions[ar].height = 18
    _merge(ws, ar, 1,  ar, 2,  '结  案', font=f_dept, align=a_center, fill=FILL_HEADER)
    _merge(ws, ar, 3,  ar, 8,  '',        font=f_normal, align=a_left)
    _merge(ws, ar, 9,  ar, 12, '品管签名：', font=f_normal, align=a_left)
    _merge(ws, ar, 13, ar, NC, '日期：',     font=f_normal, align=a_left)

    # Logo
    _logo_path = os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '..', '..', '..', 'src', 'assets', 'logo-banner.png'
    ))
    if os.path.exists(_logo_path):
        try:
            from openpyxl.drawing.image import Image as XlImage
            img = XlImage(_logo_path)
            img.width  = 110
            img.height = 30
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            _EPX = 9525   # EMU per pixel (96 DPI)
            marker = AnchorMarker(col=0, colOff=5 * _EPX, row=0, rowOff=3 * _EPX)
            img.anchor = OneCellAnchor(_from=marker,
                                       ext=XDRPositiveSize2D(cx=img.width * _EPX, cy=img.height * _EPX))
            ws.add_image(img)
        except Exception:
            pass

    _auto_col_widths(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 路由 ──────────────────────────────────────────────

@rd_bp.post('/ecr/export')
def export_ecr():
    from result import Result
    d = request.get_json() or {}
    changes = d.pop('changes', None)
    try:
        xlsx_bytes = _build_ecr_xlsx(d, changes)
    except Exception as e:
        return Result.fail(f'生成失败：{str(e)}').to_response()

    ecr_code = d.get('ecr_code', 'ECR')
    project  = d.get('project', '')
    filename = f"{ecr_code} {project} 变更申请单.xlsx".replace('/', '-')
    encoded  = urllib.parse.quote(filename)

    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}",
            'Content-Length': str(len(xlsx_bytes)),
        }
    )


def _parse_checked_text(text, opts):
    """从 ☑/☐ 标记文本中提取已选项"""
    selected = []
    for opt in opts:
        if f'☑ {opt}' in text or f'☑{opt}' in text:
            selected.append(opt)
    return selected


def _parse_ecr_rows_xlsx(ws):
    """用 openpyxl 读取 ECR xlsx 明细行，返回 (fields_dict, changes_list)"""
    def _v(r, c):
        return str(ws.cell(r, c).value or '').strip()

    dist_opts   = ['研发', '业务', '采购', '生产', '生管', '品牌', '服务', '品管']
    reason_opts = ['品质不良', '价格变动', '设计优化', '结构优化', '成本优化', '工艺优化', '其他']
    type_opts   = ['设计变更', '制程变更', '其他']

    issuing_unit   = _v(3, 2)
    date_val       = _v(3, 4)
    ecr_code       = _v(3, 7)
    project        = _v(3, 10)
    ct_text        = _v(3, 15)
    distribution   = _parse_checked_text(_v(4, 3), dist_opts)
    reason_list    = _parse_checked_text(_v(5, 3), reason_opts)
    change_reason  = reason_list[0] if reason_list else ''
    type_list      = _parse_checked_text(ct_text, type_opts)
    change_type    = type_list[0] if type_list else ''
    change_subject = _v(7, 1)
    change_desc    = _v(9, 1)

    changes = []
    for r in range(13, ws.max_row + 1):
        seq_val = ws.cell(r, 1).value
        if seq_val is None:
            break
        drawing      = str(ws.cell(r, 3).value or '').strip()
        main_drawing = str(ws.cell(r, 2).value or '').strip()
        if not drawing and not main_drawing:
            continue

        fill_rgb = ''
        try:
            fill = ws.cell(r, 1).fill
            if fill and fill.fill_type == 'solid':
                fill_rgb = fill.fgColor.rgb or ''
        except Exception:
            pass
        is_cancel_row = 'FDECEA' in fill_rgb.upper()

        change_method = str(ws.cell(r, 9).value or '').strip()
        change_kind   = str(ws.cell(r, 11).value or '').strip()
        level         = str(ws.cell(r, 4).value or '').strip()
        name          = str(ws.cell(r, 5).value or '').strip()
        spec          = str(ws.cell(r, 7).value or '').strip()

        try:
            seq = int(float(seq_val))
        except (TypeError, ValueError):
            seq = len(changes) + 1

        if is_cancel_row:
            row_type = 'cancel' if change_method == '取消' else 'deleted'
        else:
            row_type = 'add' if change_method == '新增' else 'added'

        substitution       = str(ws.cell(r, 10).value or '').strip()
        handling           = str(ws.cell(r, 17).value or '').strip()
        responsible_person = str(ws.cell(r, 18).value or '').strip()

        changes.append({
            'seq': seq, 'level': level, 'main_drawing': main_drawing,
            'drawing': drawing, 'name': name, 'spec': spec,
            'change_method': change_method, 'change_kind': change_kind,
            'substitution': substitution,
            'handling': handling, 'responsible_person': responsible_person,
            'row_type': row_type,
        })

    # 填写人员：在明细区末尾查找"填写人员：xxx"格式
    submitter = ''
    for r in range(13, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or '').strip()
        if v.startswith('填写人员：'):
            submitter = v[len('填写人员：'):].strip()
            break

    return {
        'issuing_unit': issuing_unit, 'date': date_val,
        'ecr_code': ecr_code, 'project': project,
        'change_type': change_type, 'distribution': distribution,
        'change_reason': change_reason, 'change_subject': change_subject,
        'change_desc': change_desc, 'submitter': submitter,
    }, changes


def _parse_ecr_rows_xls(ws):
    """用 xlrd 读取 ECR xls 明细行，返回 (fields_dict, changes_list)"""
    def _v(r, c):
        try:
            v = ws.cell_value(r, c)
            return str(int(v)) if isinstance(v, float) and v == int(v) else str(v or '').strip()
        except Exception:
            return ''

    dist_opts   = ['研发', '业务', '采购', '生产', '生管', '品牌', '服务', '品管']
    reason_opts = ['品质不良', '价格变动', '设计优化', '结构优化', '成本优化', '工艺优化', '其他']
    type_opts   = ['设计变更', '制程变更', '其他']

    # xls 行列从 0 开始，对应 xlsx 的行r减1、列c减1
    issuing_unit   = _v(2, 1)   # row3,col2
    date_val       = _v(2, 3)   # row3,col4
    ecr_code       = _v(2, 6)   # row3,col7
    project        = _v(2, 9)   # row3,col10
    ct_text        = _v(2, 14)  # row3,col15
    distribution   = _parse_checked_text(_v(3, 2), dist_opts)
    reason_list    = _parse_checked_text(_v(4, 2), reason_opts)
    change_reason  = reason_list[0] if reason_list else ''
    type_list      = _parse_checked_text(ct_text, type_opts)
    change_type    = type_list[0] if type_list else ''
    change_subject = _v(6, 0)   # row7,col1
    change_desc    = _v(8, 0)   # row9,col1

    changes = []
    for r in range(12, ws.nrows):  # row13 = index 12
        seq_raw = ws.cell_value(r, 0)
        if seq_raw == '' or seq_raw is None:
            break
        drawing      = str(ws.cell_value(r, 2) or '').strip()
        main_drawing = str(ws.cell_value(r, 1) or '').strip()
        if not drawing and not main_drawing:
            continue

        change_method      = str(ws.cell_value(r, 8)  or '').strip()
        substitution       = str(ws.cell_value(r, 9)  or '').strip()
        change_kind        = str(ws.cell_value(r, 10) or '').strip()
        level              = str(ws.cell_value(r, 3)  or '').strip()
        name               = str(ws.cell_value(r, 4)  or '').strip()
        spec               = str(ws.cell_value(r, 6)  or '').strip()
        handling           = str(ws.cell_value(r, 16) or '').strip()
        responsible_person = str(ws.cell_value(r, 17) or '').strip()

        try:
            seq = int(float(seq_raw))
        except (TypeError, ValueError):
            seq = len(changes) + 1

        # xls 无填充色信息可直接读，通过变更方式判断行类型
        row_type = 'cancel' if change_method == '取消' else (
                   'deleted' if change_method == '取消' else (
                   'add' if change_method == '新增' else 'added'))

        changes.append({
            'seq': seq, 'level': level, 'main_drawing': main_drawing,
            'drawing': drawing, 'name': name, 'spec': spec,
            'change_method': change_method, 'change_kind': change_kind,
            'substitution': substitution,
            'handling': handling, 'responsible_person': responsible_person,
            'row_type': row_type,
        })

    # 填写人员：在明细区末尾查找"填写人员：xxx"格式（xls 行索引从0开始）
    submitter = ''
    for r in range(12, ws.nrows):
        v = str(ws.cell_value(r, 0) or '').strip()
        if v.startswith('填写人员：'):
            submitter = v[len('填写人员：'):].strip()
            break

    return {
        'issuing_unit': issuing_unit, 'date': date_val,
        'ecr_code': ecr_code, 'project': project,
        'change_type': change_type, 'distribution': distribution,
        'change_reason': change_reason, 'change_subject': change_subject,
        'change_desc': change_desc, 'submitter': submitter,
    }, changes


@rd_bp.post('/ecr/parse-ecr')
def parse_ecr():
    """解析已导出的 ECR xlsx/xls，返回表单字段和变更明细。
    请求体：{ ecr_path: str }"""
    from result import Result

    d    = request.get_json() or {}
    path = (d.get('ecr_path') or '').strip()
    if not path:
        return Result.fail('请提供文件路径').to_response()
    if not os.path.exists(path):
        return Result.fail('文件不存在，请重新选择').to_response()

    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            fields, changes = _parse_ecr_rows_xls(ws)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            fields, changes = _parse_ecr_rows_xlsx(ws)
            wb.close()

        return Result.ok({**fields, 'changes': changes}).to_response()

    except Exception as e:
        return Result.fail(f'解析失败：{str(e)}').to_response()


@rd_bp.post('/ecr/export-ecn')
def export_ecn():
    from result import Result
    d = request.get_json() or {}
    changes = d.pop('changes', None)
    try:
        xlsx_bytes = _build_ecn_xlsx(d, changes)
    except Exception as e:
        return Result.fail(f'生成失败：{str(e)}').to_response()

    ecn_code = d.get('ecn_code', 'ECN')
    product  = d.get('product', d.get('project', ''))
    filename = f"{ecn_code} {product} 变更通知单.xlsx".replace('/', '-')
    encoded  = urllib.parse.quote(filename)

    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}",
            'Content-Length': str(len(xlsx_bytes)),
        }
    )


@rd_bp.post('/ecr/compare-bom')
def compare_bom():
    from result import Result
    d = request.get_json() or {}
    before_path = (d.get('bom_before_path') or '').strip()
    after_path  = (d.get('bom_after_path')  or '').strip()

    if not before_path or not after_path:
        return Result.fail('请先选择两个BOM文件').to_response()
    if not os.path.exists(before_path):
        return Result.fail('变更前文件不存在，请重新选择').to_response()
    if not os.path.exists(after_path):
        return Result.fail('变更审核中文件不存在，请重新选择').to_response()

    # 文件合法性校验（变更前文件不校验状态）
    err = _validate_bom(before_path, role='any')
    if err:
        return Result.fail(err).to_response()
    err = _validate_bom(after_path, role='after')
    if err:
        return Result.fail(err).to_response()

    try:
        result = _compare_bom(before_path, after_path)
        return Result.ok(result).to_response()
    except Exception as e:
        return Result.fail(f'比对失败：{str(e)}').to_response()


# ── 变更提醒 CRUD ──────────────────────────────────

def _check_rd_admin(request):
    """从请求头 X-User-Permissions 判断是否有 rd:admin 权限（前端传递），
    或从 X-User-Roles 判断是否为 admin 角色。返回 True / False。"""
    roles       = (request.headers.get('X-User-Roles', '') or '').split(',')
    permissions = (request.headers.get('X-User-Permissions', '') or '').split(',')
    return 'admin' in roles or 'rd:admin' in permissions


@rd_bp.get('/reminders')
def list_reminders():
    """返回所有在架（is_active=True）的变更提醒（所有 rd 用户可读）"""
    from result import Result
    from database.models.rd import EcrReminder
    items = EcrReminder.query.filter_by(is_active=True).order_by(EcrReminder.created_at.desc()).all()
    return Result.ok([i.to_dict() for i in items]).to_response()


@rd_bp.get('/reminders/all')
def list_reminders_all():
    """返回全部提醒（含下架历史），仅 rd:admin 可用"""
    from result import Result
    from database.models.rd import EcrReminder
    if not _check_rd_admin(request):
        return Result.fail('权限不足：需要研发部管理员权限').to_response()
    items = EcrReminder.query.order_by(EcrReminder.created_at.desc()).all()
    return Result.ok([i.to_dict() for i in items]).to_response()


@rd_bp.post('/reminders')
def create_reminder():
    """新建变更提醒，仅 rd:admin 可用"""
    from result import Result
    from database.base import db
    from database.models.rd import EcrReminder
    if not _check_rd_admin(request):
        return Result.fail('权限不足：需要研发部管理员权限').to_response()
    d = request.get_json() or {}
    content = (d.get('content') or '').strip()
    if not content:
        return Result.fail('提醒内容不能为空').to_response()
    item = EcrReminder(
        content    = content,
        notes      = (d.get('notes') or '').strip() or None,
        created_by = (d.get('created_by') or '').strip() or None,
    )
    db.session.add(item)
    db.session.commit()
    return Result.ok(item.to_dict()).to_response()


@rd_bp.put('/reminders/<int:rid>')
def update_reminder(rid):
    """编辑提醒内容/备注，仅 rd:admin 可用"""
    from result import Result
    from database.base import db
    from database.models.rd import EcrReminder
    if not _check_rd_admin(request):
        return Result.fail('权限不足：需要研发部管理员权限').to_response()
    item = EcrReminder.query.get(rid)
    if not item:
        return Result.fail('提醒不存在').to_response()
    d = request.get_json() or {}
    content = (d.get('content') or '').strip()
    if not content:
        return Result.fail('提醒内容不能为空').to_response()
    item.content = content
    item.notes   = (d.get('notes') or '').strip() or None
    db.session.commit()
    return Result.ok(item.to_dict()).to_response()


@rd_bp.put('/reminders/<int:rid>/deactivate')
def deactivate_reminder(rid):
    """下架（软删除）指定提醒，仅 rd:admin 可用"""
    from result import Result
    from database.base import db
    from database.models.rd import EcrReminder
    if not _check_rd_admin(request):
        return Result.fail('权限不足：需要研发部管理员权限').to_response()
    item = EcrReminder.query.get(rid)
    if not item:
        return Result.fail('提醒不存在').to_response()
    item.is_active = False
    db.session.commit()
    return Result.ok(item.to_dict()).to_response()


@rd_bp.put('/reminders/<int:rid>/activate')
def activate_reminder(rid):
    """重新上架已下架提醒，仅 rd:admin 可用"""
    from result import Result
    from database.base import db
    from database.models.rd import EcrReminder
    if not _check_rd_admin(request):
        return Result.fail('权限不足：需要研发部管理员权限').to_response()
    item = EcrReminder.query.get(rid)
    if not item:
        return Result.fail('提醒不存在').to_response()
    item.is_active = True
    db.session.commit()
    return Result.ok(item.to_dict()).to_response()


# ─────────────────────────────────────────────────────────────
# PDM 转 BOM
# ─────────────────────────────────────────────────────────────

_PTB_RESOURCES_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'resources')

# PDM 文件中需要校验非空的列名（来自 page_PTB.py list_checked_column）
_PTB_CHECKED_COLUMNS = [
    '品号', '品名', '规格', '数量', '单位', '库存单位', '品号群组', '生产工厂', '品号类型', '出入仓库',
    '默认销售域', '默认采购域', '采购域', '采购税分类', '销售域', '销售税分类', '默认发货工厂', '公司',
    '存货会计分类', '存货成本分类',
]

# 物料模板列对应的 PDM 列名（按位置与模板第6行对齐，来自 list_columns_to_material）
_PTB_COLUMNS_TO_MATERIAL = [
    '快捷码', '品号', '品名', '规格', '图号', '备注', '库存单位', '品号描述', '品号群组', '批号管理',
    '批号编码规则', '生产工厂', '品号类型', '出入仓库', '默认销售域', '默认采购域', '工作中心', '生产部门', '批号管理',
    '制造固定前置天数', '制造变动前置天数', '前置天数计算批量', '工单单据类型', '请购单单据类型', '补货政策',
    '最小补量', '补货倍量', '合并开单', '并单周期', '默认调出工厂', '协同关系',
    '采购域', '采购税分类', '采购人员', '供应策略', '最低补量', '采购倍量', '采购固定前置天数',
    '采购变动前置天数', '批量', '销售域', '销售税分类', '默认发货工厂', '公司', '存货会计分类', '存货成本分类',
]

# BOM 模板列对应的逻辑名（来自 list_columns_to_bom）
_PTB_COLUMNS_TO_BOM = [
    '工厂', '主件品号', '标准批量', '元件品号', '组成用量', '底数', '固定损耗量', '变动损耗', '工艺', '超领率',
    '缺领率', '供料方式', '插件位置',
]


def _ptb_build_erp_data(columns, table_data):
    """将 PDM 数据按物料模板列顺序提取，生成 ERP 物料导入数据"""
    result = []
    for row in table_data:
        material_item = []
        for col_name in _PTB_COLUMNS_TO_MATERIAL:
            if col_name in columns and col_name != '工作中心' and col_name != '生产部门':
                idx = columns.index(col_name)
                material_item.append(row[idx])
            elif col_name == '工单单据类型':
                material_item.append(5101)
            elif col_name == '请购单单据类型':
                material_item.append(3101)
            elif col_name == '工作中心':
                material_item.append(1001)
            elif col_name == '生产部门':
                material_item.append(100802)
            else:
                material_item.append('')
        result.append(material_item)
    return result


def _ptb_build_bom_data(columns, table_data, total_level):
    """根据 PDM 层级结构构建 BOM 数据"""
    code_idx  = columns.index('品号')  if '品号'  in columns else 0
    level_idx = columns.index('层次')  if '层次'  in columns else -1
    count_idx = columns.index('数量')  if '数量'  in columns else -1

    def get_depth(row):
        val = row[level_idx] if level_idx >= 0 else ''
        return len(str(val).split('.'))

    dict_bom = {}
    for depth in range(1, total_level + 1):
        for ri in range(len(table_data)):
            row = table_data[ri]
            if get_depth(row) != depth:
                continue
            code = row[code_idx]
            for nj in range(ri + 1, len(table_data)):
                next_row = table_data[nj]
                next_depth = get_depth(next_row)
                if next_depth <= depth:
                    break
                if next_depth == depth + 1:
                    next_code = next_row[code_idx]
                    raw_count = next_row[count_idx] if count_idx >= 0 else 1
                    try:
                        qty = int(float(raw_count)) if raw_count else 1
                    except (ValueError, TypeError):
                        qty = 1
                    bom_row = []
                    for col_name in _PTB_COLUMNS_TO_BOM:
                        if col_name == '工厂':
                            bom_row.append(100)
                        elif col_name == '主件品号':
                            bom_row.append(code)
                        elif col_name == '标准批量':
                            bom_row.append(1)
                        elif col_name == '元件品号':
                            bom_row.append(next_code)
                        elif col_name == '组成用量':
                            bom_row.append(qty)
                        elif col_name == '底数':
                            bom_row.append(1)
                        elif col_name == '供料方式':
                            bom_row.append(1)
                        else:
                            bom_row.append('')
                    dict_bom.setdefault(code, []).append(bom_row)
    result = []
    for key in dict_bom:
        result.extend(dict_bom[key])
    return result


@rd_bp.post('/pdm2bom/process')
def pdm2bom_process():
    """解析 PDM 导出文件，校验必填列，返回表格数据与错误索引"""
    from result import Result
    import openpyxl as _xl

    d = request.get_json() or {}
    file_path = (d.get('file_path') or '').strip()
    if not file_path or not os.path.isfile(file_path):
        return Result.fail('文件不存在或路径无效').to_response()
    if not file_path.lower().endswith('.xlsx'):
        return Result.fail('仅支持 .xlsx 格式').to_response()

    try:
        wb = _xl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return Result.fail(f'文件读取失败：{e}').to_response()

    if not all_rows:
        return Result.fail('文件为空').to_response()

    # 第一行为列名
    columns = [str(c) if c is not None else '' for c in all_rows[0]]

    # 校验关键列存在
    if '品号' not in columns or '层次' not in columns:
        return Result.fail('文件格式不符：缺少"品号"或"层次"列').to_response()

    code_idx  = columns.index('品号')
    level_idx = columns.index('层次')
    count_idx = columns.index('数量') if '数量' in columns else -1

    # 过滤数据行：跳过空行、含'.'的品号、14ST10前缀
    table_data = []
    total_level = 0
    for raw_row in all_rows[1:]:
        if not any(v for v in raw_row if v is not None):
            continue
        code = str(raw_row[code_idx]) if raw_row[code_idx] is not None else ''
        if '.' in code or code[:6] == '14ST10':
            continue
        level_val = str(raw_row[level_idx]) if raw_row[level_idx] is not None else ''
        depth = len(level_val.split('.'))
        if depth > total_level:
            total_level = depth
        table_data.append([str(v) if v is not None else '' for v in raw_row])

    # 必填列索引
    required_col_indices = [columns.index(c) for c in _PTB_CHECKED_COLUMNS if c in columns]

    # 逐行校验，记录缺失列
    error_map = {}
    for ri, row in enumerate(table_data):
        missing = [ci for ci in required_col_indices if not row[ci]]
        if missing:
            error_map[str(ri)] = missing

    return Result.ok({
        'columns':             columns,
        'table_data':          table_data,
        'required_col_indices': required_col_indices,
        'error_map':           error_map,
        'total_level':         total_level,
    }).to_response()


@rd_bp.post('/pdm2bom/export-erp')
def pdm2bom_export_erp():
    """生成 ERP 物料导入 xlsx，返回 arraybuffer"""
    from result import Result
    import openpyxl as _xl

    d = request.get_json() or {}
    columns    = d.get('columns')    or []
    table_data = d.get('table_data') or []
    if not columns or not table_data:
        return Result.fail('数据不能为空').to_response()

    erp_data   = _ptb_build_erp_data(columns, table_data)
    code_idx   = columns.index('品号') if '品号' in columns else 0
    first_code = table_data[0][code_idx] if table_data else 'ERP'

    template_path = os.path.join(_PTB_RESOURCES_DIR, 'template_material.xlsx')
    try:
        wb = _xl.load_workbook(template_path)
    except Exception as e:
        return Result.fail(f'物料模板加载失败：{e}').to_response()

    ws = wb.active
    for i, row in enumerate(erp_data, start=6):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()
    filename = f'ERP-{first_code}.xlsx'
    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}",
            'Content-Length': str(len(xlsx_bytes)),
        },
    )


@rd_bp.post('/pdm2bom/export-bom')
def pdm2bom_export_bom():
    """生成 BOM 导入 xlsx，返回 arraybuffer"""
    from result import Result
    import openpyxl as _xl

    d = request.get_json() or {}
    columns     = d.get('columns')     or []
    table_data  = d.get('table_data')  or []
    total_level = d.get('total_level') or 0
    if not columns or not table_data:
        return Result.fail('数据不能为空').to_response()

    bom_data   = _ptb_build_bom_data(columns, table_data, total_level)
    code_idx   = columns.index('品号') if '品号' in columns else 0
    first_code = table_data[0][code_idx] if table_data else 'BOM'

    template_path = os.path.join(_PTB_RESOURCES_DIR, 'template_bom.xlsx')
    try:
        wb = _xl.load_workbook(template_path)
    except Exception as e:
        return Result.fail(f'BOM 模板加载失败：{e}').to_response()

    ws = wb.active
    for i, row in enumerate(bom_data, start=6):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()
    filename = f'BOM-{first_code}.xlsx'
    return Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}",
            'Content-Length': str(len(xlsx_bytes)),
        },
    )
